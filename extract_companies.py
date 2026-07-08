import openpyxl
import json

wb = openpyxl.load_workbook('CRM_Enrichment_Output.xlsx')
ws = wb['ENRICHMENT_READY']

companies = []
for row in range(2, ws.max_row+1):
    mid = ws.cell(row, 1).value or ''
    std_name = ws.cell(row, 3).value or ''
    orig_name = ws.cell(row, 2).value or ''
    email = ws.cell(row, 22).value or ''
    phone = ws.cell(row, 23).value or ''
    contact = ws.cell(row, 21).value or ''
    ws_q1 = ws.cell(row, 27).value or ''
    ws_q2 = ws.cell(row, 28).value or ''
    li_q1 = ws.cell(row, 31).value or ''
    li_q2 = ws.cell(row, 32).value or ''
    companies.append({
        'mid': mid, 'std_name': std_name, 'orig_name': orig_name,
        'email': email, 'phone': phone, 'contact': contact,
        'ws_q1': ws_q1, 'ws_q2': ws_q2, 'li_q1': li_q1, 'li_q2': li_q2
    })

# Write to JSON for processing
with open('companies.json', 'w', encoding='utf-8') as f:
    json.dump(companies, f, ensure_ascii=False, indent=2)

print(f'Total companies: {len(companies)}')
print('First 10 company names:')
for i, c in enumerate(companies[:10]):
    print(f'  {c["mid"]}: {c["std_name"][:60]}')
