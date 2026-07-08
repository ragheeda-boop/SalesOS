import json, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

wb = openpyxl.Workbook()
ws1 = wb.active
ws1.title = "ENRICHED_COMPANIES"

headers = [
    "MID", "COMPANY_NAME", "WEBSITE", "WEBSITE_CONFIDENCE",
    "LINKEDIN_URL", "LINKEDIN_CONFIDENCE", "SOURCE_1", "SOURCE_2",
    "HAS_WEBSITE", "HAS_LINKEDIN", "ENRICHMENT_DATE"
]

header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

for col, h in enumerate(headers, 1):
    cell = ws1.cell(row=1, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border
    cell.alignment = Alignment(horizontal='center')

ws1.freeze_panes = 'A2'

ws2 = wb.create_sheet("FAILED_LOOKUPS")
failed_headers = [
    "MID", "COMPANY_NAME", "WEBSITE_SEARCHED", "LINKEDIN_SEARCHED",
    "FAILURE_REASON", "ENRICHMENT_DATE"
]

for col, h in enumerate(failed_headers, 1):
    cell = ws2.cell(row=1, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border
    cell.alignment = Alignment(horizontal='center')

ws2.freeze_panes = 'A2'

# Adjust column widths
for ws in [ws1, ws2]:
    for col in ws.columns:
        max_len = 0
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except:
                pass
        adjusted = min(max_len + 2, 60)
        ws.column_dimensions[col_letter].width = max(adjusted, 15)

output_path = r'C:\Users\raghe\OneDrive - RATL Technology Ltd\Muhide\CRM_Enriched_Final.xlsx'
wb.save(output_path)
print('Schema created: ' + output_path)
