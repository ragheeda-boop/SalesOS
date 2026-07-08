import openpyxl, sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

wb = openpyxl.load_workbook('Sales_Intelligence_Output.xlsx')

# ICP_PRIORITY_ACCOUNTS
ws = wb['ICP_PRIORITY_ACCOUNTS']
headers = {}
for c in range(1, ws.max_column + 1):
    headers[ws.cell(1, c).value] = c

print("=== ICP_PRIORITY_ACCOUNTS (Tier 1) ===")
print(f"Total: {ws.max_row - 1}")
hws = 0; hli = 0; needs_ws = 0; needs_li = 0; needs_both = 0

for r in range(2, ws.max_row + 1):
    ws_url = str(ws.cell(r, headers.get('Website', 7)).value or '')
    li_url = str(ws.cell(r, headers.get('LinkedIn URL', 8)).value or '')
    name = str(ws.cell(r, 2).value or '')[:40]
    mid = ws.cell(r, 1).value
    site_ok = bool(ws_url and ws_url != 'None' and ws_url.startswith('http'))
    li_ok = bool(li_url and li_url != 'None' and 'linkedin.com' in li_url)
    if site_ok: hws += 1
    if li_ok: hli += 1
    if not site_ok: needs_ws += 1
    if not li_ok: needs_li += 1
    if not site_ok and not li_ok: needs_both += 1
    print(f"  {mid}: {name} | WS={'YES' if site_ok else 'NO'} | LI={'YES' if li_ok else 'NO'}")

tot = ws.max_row - 1
print(f"\nCoverage (Tier 1):")
print(f"  Website: {hws}/{tot} = {hws/tot*100:.1f}%")
print(f"  LinkedIn: {hli}/{tot} = {hli/tot*100:.1f}%")
print(f"  Need website: {needs_ws}, Need LinkedIn: {needs_li}, Need both: {needs_both}")

# All 236
ws2 = wb['SALES_INTELLIGENCE_MASTER']
h2 = {}
for c in range(1, ws2.max_column + 1):
    h2[ws2.cell(1, c).value] = c

print(f"\n=== SALES_INTELLIGENCE_MASTER (all 236) ===")
all_ws = 0; all_li = 0; all_tot = ws2.max_row - 1
for r in range(2, ws2.max_row + 1):
    ws_url = str(ws2.cell(r, h2.get('Validated Website', 9)).value or '')
    li_url = str(ws2.cell(r, h2.get('LinkedIn URL', 12)).value or '')
    if ws_url and ws_url != 'None' and ws_url.startswith('http'): all_ws += 1
    if li_url and li_url != 'None' and 'linkedin.com' in li_url: all_li += 1

print(f"  Total: {all_tot}")
print(f"  Website: {all_ws}/{all_tot} = {all_ws/all_tot*100:.1f}%")
print(f"  LinkedIn: {all_li}/{all_tot} = {all_li/all_tot*100:.1f}%")
print(f"  Need website: {all_tot - all_ws}, Need LinkedIn: {all_tot - all_li}")
