"""
Build comprehensive Tier 1 LinkedIn Discovery output.
"""
import openpyxl, json, datetime
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

TODAY = datetime.datetime.now().strftime('%Y-%m-%d')
OUT = 'C:\\Users\\raghe\\OneDrive - RATL Technology Ltd\\Muhide\\Tier1_LinkedIn_Discovery.xlsx'
SRC = 'C:\\Users\\raghe\\OneDrive - RATL Technology Ltd\\Muhide\\Sales_Intelligence_Output_COPY.xlsx'

wb = openpyxl.Workbook()

# ===================== NEWLY VERIFIED =====================
ws_verified = wb.active
ws_verified.title = 'LINKEDIN_VERIFIED'

headers = [
    'Company ID', 'Company Name', 'Arabic Name', 'ICP Score',
    'LinkedIn URL', 'Followers', 'Employees', 'Industry',
    'Headquarters', 'Founded', 'Type', 'Website',
    'About', 'Confidence', 'Discovery Method', 'Discovered Date'
]
hdr_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
hdr_font = Font(bold=True, color="FFFFFF", size=11)
thin = Border(left=Side(style='thin'), right=Side(style='thin'),
              top=Side(style='thin'), bottom=Side(style='thin'))

for ci, h in enumerate(headers, 1):
    c = ws_verified.cell(1, ci, h)
    c.fill = hdr_fill
    c.font = hdr_font
    c.alignment = Alignment(horizontal='center', wrap_text=True)

verified_data = [
    ['COMP-000068', 'Whole Foods Trading', 'شركة الاغذية الكاملة للتجارة', 'Tier 1',
     'https://www.linkedin.com/company/wholefoodstrading',
     996, 6, 'Food and Beverage Services',
     'Jeddah, Makkah', '', 'Self Owned', 'http://wholefoods.sa',
     'Whole Foods Trading Company is operating in Jeddah, Saudi Arabia, for over 20 years. Agricultural commodity distributor.',
     100, 'Web Search (site:linkedin.com/company)', TODAY],
    ['COMP-000216', 'Saudi Contractors Authority', 'الهيئة السعودية للمقاولين', 'Tier 1',
     'https://www.linkedin.com/company/sca2030',
     97118, 172, 'Non-profit Organizations',
     'Riyadh, Saudi Arabia', '2016', 'Government Agency', 'http://www.sca.sa',
     'Official LinkedIn of SCA. Established by Council of Ministers Resolution No. 510. Develops the contracting sector in KSA.',
     100, 'Web Search (site:linkedin.com/company)', TODAY],
    ['COMP-000308', 'PROGMED', 'شركة مقياس التقدم الطبية', 'Tier 1',
     'https://www.linkedin.com/company/progmed',
     '', '11-50', 'Medical Equipment Manufacturing',
     'Jeddah', '2016', 'Partnership', 'http://www.progmed.sa',
     'Leading manufacturer of high-quality medical supplies. World-class medical supplies manufacturer in Jeddah, KSA.',
     100, 'Website Footer (progmed.sa)', TODAY],
    ['COMP-000455', 'GT Medical', '', 'Tier 1',
     'https://www.linkedin.com/company/gtmedical',
     41411, '51-200', 'Medical Equipment Manufacturing',
     'Jeddah', '2005', 'Privately Held', 'http://www.gtmedical.com',
     'Leading Saudi MedTech & Pharma company. Focus on advancing humanity by providing innovative health solutions in KSA.',
     100, 'Website Footer (gtmedical.com)', TODAY],
]

green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
for ri, row in enumerate(verified_data, 2):
    for ci, val in enumerate(row):
        c = ws_verified.cell(ri, ci + 1, val)
        c.border = thin
        if val and val == 100: c.fill = green

# ===================== PREVIOUSLY VERIFIED =====================
ws_prev = wb.create_sheet('LINKEDIN_PREVIOUS')
for ci, h in enumerate(headers, 1):
    c = ws_prev.cell(1, ci, h)
    c.fill = hdr_fill
    c.font = hdr_font
    c.alignment = Alignment(horizontal='center', wrap_text=True)

prev_data = [
    ['COMP-000069', 'Al-Taawin Medical Group', 'شركة التعاون الطبية', 'Tier 1',
     'https://www.linkedin.com/company/altaawin-medical-group',
     1128, '', 'Medical', '', '', '', '',
     'Medical group in Saudi Arabia', 85, 'Previous pipeline', ''],
    ['COMP-000084', 'RATL Technology', 'شركة رتل التقنية المحدودة', 'Tier 1',
     'https://sa.linkedin.com/company/ratl-technology',
     '', '', 'Technology', '', '', '', '',
     'Technology company in Saudi Arabia', 70, 'Previous pipeline', ''],
    ['COMP-000133', 'MEGAMIND IT Solutions', '', 'Tier 1',
     'https://sa.linkedin.com/company/megaminditsolutions',
     '', '51-100', 'IT Services', '', '', '', '',
     'CEO: Hamza Batterjee. IT solutions in Saudi Arabia.', 100, 'Previous pipeline', ''],
    ['COMP-000718', 'LaseTech KSA', '', 'Tier 1',
     'https://www.linkedin.com/company/lasetech-ksa',
     297, 14, 'Technology', '', '', '', '',
     'Founder: Islam Mostafa. Technology solutions in Saudi Arabia.', 100, 'Previous pipeline', ''],
]

for ri, row in enumerate(prev_data, 2):
    for ci, val in enumerate(row):
        c = ws_prev.cell(ri, ci + 1, val)
        c.border = thin

# ===================== REVIEW REQUIRED =====================
ws_review = wb.create_sheet('LINKEDIN_REVIEW_REQUIRED')
for ci, h in enumerate(['Company ID', 'Company Name', 'Issue', 'Found URL', 'Confidence', 'Notes'], 1):
    c = ws_review.cell(1, ci, h)
    c.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    c.font = Font(bold=True)

review_data = [
    ['COMP-000313', 'BrandIT / العلامة المميزة', 'Website has LinkedIn icon but matched Turkish company',
     'https://www.linkedin.com/company/brand-it', 30,
     'brandit.sa website footer shows LinkedIn social media link. Direct URL resolution found Turkish Brandit (Istanbul, Advertising). Saudi brandit.sa LinkedIn page may use different slug. Needs manual verification.'],
    ['COMP-000514', 'ETE Group / الشرقية للخرسانة', 'LinkedIn page found but company mismatch',
     'https://www.linkedin.com/company/ete-group', 10,
     'Page found is a Public Safety company (2-10 employees), not the Saudi precast concrete manufacturer. Website ete-group.com returned empty page.'],
]

for ri, row in enumerate(review_data, 2):
    for ci, val in enumerate(row):
        c = ws_review.cell(ri, ci + 1, val)
        c.border = thin

# ===================== NOT FOUND =====================
ws_notfound = wb.create_sheet('LINKEDIN_NOT_FOUND')
for ci, h in enumerate(['Company ID', 'Company Name', 'Arabic Name', 'Website', 'Search Methods Used', 'Notes'], 1):
    c = ws_notfound.cell(1, ci, h)
    c.fill = PatternFill(start_color="FF4444", end_color="FF4444", fill_type="solid")
    c.font = Font(bold=True, color="FFFFFF")

not_found = [
    ['COMP-000032', 'مصنع البركة للحلويات والأغذية', 'Al Baraka Sweets', 'https://www.babader.com',
     'site:linkedin, web search, Arabic search', 'Employee profiles found (مروان عبده سيف الذبحاني) but no company page. Jeddah-based food factory.'],
    ['COMP-000056', 'شركه منيف بن عامر النهدي وشركاه للمقاولات', 'Munif Al Nahdi Contracting', 'https://www.nahdico.com',
     'site:linkedin, web search, Arabic search', 'No LinkedIn presence found. Saudi contracting company.'],
    ['COMP-000058', 'شركة اوغاريت كو للمواد الغذائية', 'Ugarit Co Foodstuff', 'https://www.ogaretco.com',
     'site:linkedin, web search, Arabic search', 'No LinkedIn page. Operates in KSA, UAE, Turkey, Germany. D&B profile exists.'],
    ['COMP-000073', 'مصنع شركة قبة الدمام للبلاستيك', 'Dammam Dome Plastic', '',
     'site:linkedin, web search', 'Directory listing on icuae.ma found. No LinkedIn or website.'],
    ['COMP-000081', 'شركة الهمة العقارية', 'Alhimma Real Estate', 'https://www.himhre.com',
     'site:linkedin, web search', 'No LinkedIn presence. Saudi real estate company.'],
    ['COMP-000098', 'شركة تقنية المعمار السريع للمقاولات', 'Speetech', 'https://www.speetech.net',
     'site:linkedin, web search', 'No LinkedIn presence. Contracting company in Saudi Arabia.'],
    ['COMP-000185', 'شركه ادمه شمران للتموين المحدوده', 'Adma Shamran Trading', 'https://www.admashamran.com',
     'site:linkedin, web search', 'Employees found (Abdul Hameed, Faisal Khalil) but no company page. Food & catering company.'],
    ['COMP-000293', 'المصنع السعودي البيئي للمنظفات', 'Saudi Env. Solvents', 'https://www.gogreensolvents.com',
     'site:linkedin, web search, Arabic search', 'No LinkedIn page. Directory listing on dalelglobal.com. Website active.'],
    ['COMP-000300', 'شركة مصفوفات البرمجية للاتصالات', 'Masfufat Software', 'https://www.masfufat.com',
     'site:linkedin, web search, Arabic search, website check', 'Salla-based e-commerce platform. No LinkedIn.'],
    ['COMP-000307', 'مصنع عاشق البن للصناعات الغذائية', 'Sout Roastery', 'https://www.soutroastery.com',
     'site:linkedin, web search, Arabic search, website check', 'Salla-based coffee roastery. No LinkedIn.'],
    ['COMP-000311', 'مصنع ليات العايد', 'Al Ayed Factory', 'https://www.ayed-sa.com',
     'site:linkedin, web search, website check', 'No LinkedIn presence. Minimal website.'],
    ['COMP-000318', 'شركة عالم سجال للخدمات الطبيه', 'Sijal Medical World', '',
     'site:linkedin, web search, Arabic search', 'No website, no LinkedIn. Medical services company.'],
    ['COMP-000322', 'شركة افتاب الغذائية المحدودة', 'Aftab Foods', '',
     'site:linkedin, web search, Arabic search', 'No website (dns failed), no LinkedIn. Food company.'],
    ['COMP-000335', 'National Center for Non-Profit Sector Development', '', 'https://www.ncnp.gov.sa',
     'site:linkedin, web search, direct URL', 'No LinkedIn page found. Government entity (website DNS failed).'],
    ['COMP-000472', 'شركة معايير البن الصناعية', 'Industrial Coffee Standards', 'https://www.ashcafes.com',
     'site:linkedin, web search, direct URL', 'Operates ASH Coffee brand (7+ branches, Jeddah). No LinkedIn. Website has Instagram/WhatsApp only.'],
    ['COMP-000486', 'مصنع الشرقية لسحب الالمنيوم', 'Eastern Aluminum Extrusion', 'https://www.easternextrusion.com',
     'site:linkedin, web search, direct URL', 'Dammam-based factory, 40K tons/yr capacity. No LinkedIn. Website active.'],
    ['COMP-000494', 'وردة الحلويات', 'Warda Sweets', '',
     'site:linkedin, web search, Arabic search', 'No website, no LinkedIn. Sweets company.'],
    ['COMP-000514', 'مصنع الشرقية للخرسانة المسبقة الصنع', 'Eastern Precast Concrete', 'https://www.ete-group.com',
     'site:linkedin, direct URL, website check', 'ETE found but wrong company (Public Safety). Saudi precast concrete likely has no LinkedIn.'],
    ['COMP-000709', 'شركة مصنع المواد العازلة', 'Insulation Materials Factory', 'https://www.sppc.com.sa',
     'site:linkedin, direct URL, website check', 'No LinkedIn page found. Insulation materials manufacturer.'],
    ['COMP-000710', 'مصنع شركة ابناء صالح محمد حماده للمخبوزات', 'Mehran Bakery', 'https://www.mehran.com.sa',
     'site:linkedin, direct URL, website check', 'No LinkedIn page found. Bakery products factory.'],
]

for ri, row in enumerate(not_found, 2):
    for ci, val in enumerate(row):
        ws_notfound.cell(ri, ci + 1, val).border = thin

# ===================== EXECUTIVE SUMMARY =====================
ws_es = wb.create_sheet('EXECUTIVE_SUMMARY')
summary = [
    ['TIER 1 LINKEDIN DISCOVERY REPORT'],
    [''],
    ['Generated:', TODAY],
    ['Source:', 'Sales_Intelligence_Output.xlsx (ICP_PRIORITY_ACCOUNTS)'],
    [''],
    ['=== COVERAGE RESULTS ==='],
    ['Metric', 'Count', '%'],
    ['Total Tier 1 Accounts', '29', '100%'],
    ['LinkedIn VERIFIED (New)', '4', '13.8%'],
    ['LinkedIn VERIFIED (Previous)', '4', '13.8%'],
    ['Total LinkedIn Coverage', '8', '27.6%'],
    ['LinkedIN Not Found', '21', '72.4%'],
    ['LinkedIn Needs Review', '2', '6.9%'],
    [''],
    ['=== NEWLY VERIFIED ==='],
    ['Company', 'LinkedIn URL', 'Followers'],
    ['Whole Foods Trading', 'linkedin.com/company/wholefoodstrading', '996'],
    ['Saudi Contractors Authority', 'linkedin.com/company/sca2030', '97,118'],
    ['PROGMED', 'linkedin.com/company/progmed', '-'],
    ['GT Medical', 'linkedin.com/company/gtmedical', '41,411'],
    [''],
    ['=== METHOD USED ==='],
    ['Method', 'Found'],
    ['Method 1: site:linkedin.com/company', '3 (Whole Foods, SCA, PROGMED)'],
    ['Method 4: Website footer/check', '1 (GT Medical)'],
    [''],
    ['=== KEY INSIGHTS ==='],
    ['1. LinkedIn coverage for Saudi SMEs remains low - 21/29 Tier 1 accounts have NO LinkedIn company page.'],
    ['2. Most companies are family-owned SMEs without digital corporate presence beyond basic websites.'],
    ['3. Companies with LinkedIn (8/29) are either larger entities (GT Medical: 51-200 emp) or government (SCA: 172 emp).'],
    ['4. 4 accounts have NO website at all, making LinkedIn discovery nearly impossible via automated means.'],
    ['5. 2 accounts need manual review (BrandIT ambiguous, ETE GROUP wrong match).'],
    [''],
    ['=== RECOMMENDATIONS ==='],
    ['1. Create LinkedIn pages for the 21 Tier 1 accounts that are missing them (high ROI for sales enablement).'],
    ['2. Manual review of COMP-000313 (BrandIT) - check website footer for exact LinkedIn URL.'],
    ['3. For accounts with employees found but no company page (COMP-000032, COMP-000185):'],
    ['   - Reach out to discovered employees via LinkedIn InMail for company introduction.'],
    ['4. Consider using Saudi business directories (muqawil, maolaty, marefa) as alternative data sources.'],
    ['5. Target 50% coverage by creating LinkedIn pages for remaining accounts.'],
]

title_font = Font(bold=True, size=14, color="1F4E79")
section_font = Font(bold=True, size=12, color="2F5496")

for ri, row in enumerate(summary, 1):
    for ci, val in enumerate(row):
        cell = ws_es.cell(ri, ci + 1, val)
        if ri == 1: cell.font = title_font
        if val and str(val).startswith('==='): cell.font = section_font

# Column widths
for ws in wb.worksheets:
    for col in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col)].width = 35 if col < 5 else 25

ws_verified.column_dimensions['A'].width = 15
ws_verified.column_dimensions['E'].width = 45
ws_verified.column_dimensions['M'].width = 60

wb.save(OUT)
print(f"Saved: {OUT}")
print(f"\nTIER 1 LINKEDIN COVERAGE:")
print(f"  Previously: 4/29 (13.8%)")
print(f"  Now:        8/29 (27.6%)")
print(f"  New finds:  4 (Whole Foods, SCA, PROGMED, GT Medical)")
