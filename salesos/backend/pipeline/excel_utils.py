"""Shared Excel utilities for data pipeline scripts.

Usage:
    from pipeline.excel_utils import style_header, auto_width, GREEN_FILL, RED_FILL, YELLOW_FILL, HDR_FONT, HDR_FILL
"""

from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

# ── Shared styles ────────────────────────────────────────────
HDR_FONT = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
HDR_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")

GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

LIGHT_BLUE_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def style_header(ws, num_cols: int, row: int = 1) -> None:
    """Apply header font, fill, border, and center alignment to a header row."""
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER


def auto_width(ws, max_width: int = 40, min_width: int = 8) -> None:
    """Auto-fit column widths based on content."""
    for col_cells in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col_cells), default=0)
        width = min(max(max_len + 2, min_width), max_width)
        ws.column_dimensions[col_cells[0].column_letter].width = width


def apply_border(ws, min_row: int, max_row: int, min_col: int, max_col: int) -> None:
    """Apply thin border to a range of cells."""
    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            ws.cell(row=row, column=col).border = THIN_BORDER
