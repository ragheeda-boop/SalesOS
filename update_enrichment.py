import json, openpyxl
from datetime import datetime
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

companies = json.load(open(r'C:\Users\raghe\OneDrive - RATL Technology Ltd\Muhide\companies.json', encoding='utf-8'))
today = datetime.now().strftime('%Y-%m-%d')

# Manually verified enrichments from web research
verified = {
    'COMP-000001': {'website': '', 'ws_conf': 0, 'linkedin': '', 'li_conf': 0, 'source1': 'Facebook: Basma Zad|dlilsa.com listing', 'source2': 'Social media (TikTok, Instagram) found but no official site'},
    'COMP-000002': {'website': '', 'ws_conf': 0, 'linkedin': '', 'li_conf': 0, 'source1': 'No direct web match', 'source2': 'Generic name - needs Arabic search refinement'},
    'COMP-000003': {'website': 'https://mlhamet-alfakher.com', 'ws_conf': 30, 'linkedin': '', 'li_conf': 0, 'source1': 'Facebook: معمل الفاخر', 'source2': 'Low confidence website match'},
    'COMP-000004': {'website': '', 'ws_conf': 0, 'linkedin': '', 'li_conf': 0, 'source1': 'Search returned مخابز بدر (Badr Bakeries)', 'source2': 'Name variant - needs manual verification'},
    'COMP-000005': {'website': '', 'ws_conf': 0, 'linkedin': '', 'li_conf': 0, 'source1': '', 'source2': ''},
    'COMP-000006': {'website': '', 'ws_conf': 0, 'linkedin': '', 'li_conf': 0, 'source1': '', 'source2': ''},
    'COMP-000007': {'website': '', 'ws_conf': 0, 'linkedin': '', 'li_conf': 0, 'source1': 'Instagram: marja_juices|Al Qatif juice restaurant', 'source2': 'No official website found'},
    'COMP-000008': {'website': '', 'ws_conf': 0, 'linkedin': 'https://www.linkedin.com/in/مؤسسة-محمص-للتجارة-والتجزئه-a93b97377', 'li_conf': 50, 'source1': 'LinkedIn personal profile', 'source2': 'No official website found'},
    'COMP-000009': {'website': '', 'ws_conf': 0, 'linkedin': 'https://www.linkedin.com/company/bahasan-trading', 'li_conf': 70, 'source1': 'LinkedIn: شركة باحسن التجارية (75 followers)', 'source2': 'directoryksa.com listing, 50+ years in food supply'},
    'COMP-000010': {'website': '', 'ws_conf': 0, 'linkedin': '', 'li_conf': 0, 'source1': '', 'source2': ''},
    'COMP-000011': {'website': '', 'ws_conf': 0, 'linkedin': '', 'li_conf': 0, 'source1': '', 'source2': ''},
    'COMP-000012': {'website': '', 'ws_conf': 0, 'linkedin': '', 'li_conf': 0, 'source1': '', 'source2': 'Generic name - needs manual verification'},
    'COMP-000033': {'website': 'https://nourmassah.com', 'ws_conf': 100, 'linkedin': 'https://sa.linkedin.com/company/creativeteamforce', 'li_conf': 85, 'source1': 'Signage manufacturer since 1989, Riyadh', 'source2': 'LinkedIn as Nour Massah Company for Industry (463 followers)'},
    'COMP-000035': {'website': '', 'ws_conf': 0, 'linkedin': '', 'li_conf': 0, 'source1': 'Email domain marveltravel.com.sa does not resolve', 'source2': ''},
    'COMP-000040': {'website': 'https://rafid.com.sa', 'ws_conf': 100, 'linkedin': '', 'li_conf': 0, 'source1': 'Rafid Logistics - automated logistics platform, Saudi', 'source2': 'Email rafid@rafid.com.sa matches domain'},
    'COMP-000043': {'website': 'https://site.mohrkeyapp.com', 'ws_conf': 100, 'linkedin': 'https://sa.linkedin.com/company/mohrkey', 'li_conf': 100, 'source1': 'Mohrkey - leading Saudi car services company', 'source2': 'LinkedIn: MOHRKEY - محركي (2,204 followers)'},
    'COMP-000045': {'website': 'http://yasminefood.com', 'ws_conf': 40, 'linkedin': '', 'li_conf': 0, 'source1': 'Domain yasminefood.com exists but default page only', 'source2': ''},
    'COMP-000050': {'website': 'https://www.7eleventrading.org', 'ws_conf': 100, 'linkedin': 'https://sa.linkedin.com/company/seven-eleven-trading', 'li_conf': 90, 'source1': 'Food distribution company, Saudi Arabia', 'source2': 'LinkedIn: Seven Eleven Trading'},
    'COMP-000025': {'website': '', 'ws_conf': 0, 'linkedin': 'https://www.linkedin.com/company/nama-arabia-trading-company', 'li_conf': 85, 'source1': 'LinkedIn: Nama Arabia Trading Company (3,188 followers)', 'source2': 'Email yhaddad@alnahdico.com - domain differs from name'},
    'COMP-000029': {'website': 'https://sedres.com', 'ws_conf': 100, 'linkedin': 'https://www.linkedin.com/company/sedres-maritime-co-ltd', 'li_conf': 100, 'source1': 'sedres.com - maritime services, Saudi Arabia', 'source2': 'LinkedIn: Sedres Maritime Co. Ltd.'},
    'COMP-000062': {'website': 'http://zincoint.com', 'ws_conf': 100, 'linkedin': 'https://sa.linkedin.com/company/zincoint', 'li_conf': 100, 'source1': 'zincoint.com - Zinco International Trading (18,732 followers)', 'source2': 'Email/domain verified: @zincoint.com'},
    'COMP-000065': {'website': 'https://fazco.co', 'ws_conf': 100, 'linkedin': '', 'li_conf': 0, 'source1': 'fazco.co - Fazco Food Supply Company', 'source2': 'Email/domain verified: @fazco.co'},
    'COMP-000066': {'website': 'https://cs-con.co', 'ws_conf': 100, 'linkedin': '', 'li_conf': 0, 'source1': 'cs-con.co - Esnad Al-Emar for Contracting', 'source2': 'Email/domain verified: @cs-con.co'},
}

def extract_domain(email):
    if not email: return ''
    import re
    m = re.search(r'@([\w.-]+)', email)
    return m.group(1).lower() if m else ''

non_personal_domains = {
    'gmail.com', 'hotmail.com', 'yahoo.com', 'outlook.com',
    'live.com', 'mail.com', 'msn.com'
}

wb = openpyxl.load_workbook(r'C:\Users\raghe\OneDrive - RATL Technology Ltd\Muhide\CRM_Enriched_Final.xlsx')
ws1 = wb['ENRICHED_COMPANIES']
ws2 = wb['FAILED_LOOKUPS']

header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

# Clear existing data rows
ws1.delete_rows(2, ws1.max_row)
ws2.delete_rows(2, ws2.max_row)

row1 = 2
row2 = 2

for c in companies:
    mid = c['mid']
    name = c['std_name']
    email = c.get('email', '')
    domain = extract_domain(email)
    
    v = verified.get(mid, {})
    
    # Determine website
    ws_url = v.get('website', '')
    ws_conf = v.get('ws_conf', 0)
    
    if not ws_url and domain and domain not in non_personal_domains:
        # Don't derive from domain if verified entry explicitly has no website
        if not mid in verified or v.get('ws_conf', 0) != 0:
            ws_url = 'https://www.' + domain
            ws_conf = max(ws_conf, 85)
    
    # Determine LinkedIn
    li_url = v.get('linkedin', '')
    li_conf = v.get('li_conf', 0)
    
    source1 = v.get('source1', '')
    source2 = v.get('source2', '')
    if not source1 and domain:
        source1 = 'Derived from email domain: ' + domain
    
    has_ws = 'Yes' if ws_url and ws_conf >= 70 else 'No'
    has_li = 'Yes' if li_url and li_conf >= 70 else 'No'
    
    # Enrichment confidence assessment
    if ws_conf >= 70 or li_conf >= 70:
        # Enriched successfully
        ws1.cell(row=row1, column=1, value=mid)
        ws1.cell(row=row1, column=2, value=name)
        ws1.cell(row=row1, column=3, value=ws_url)
        ws1.cell(row=row1, column=4, value=ws_conf)
        ws1.cell(row=row1, column=5, value=li_url)
        ws1.cell(row=row1, column=6, value=li_conf)
        ws1.cell(row=row1, column=7, value=source1)
        ws1.cell(row=row1, column=8, value=source2)
        ws1.cell(row=row1, column=9, value=has_ws)
        ws1.cell(row=row1, column=10, value=has_li)
        ws1.cell(row=row1, column=11, value=today)
        for col in range(1, 12):
            ws1.cell(row=row1, column=col).border = thin_border
        row1 += 1
    else:
        # Failed enrichment
        ws2.cell(row=row2, column=1, value=mid)
        ws2.cell(row=row2, column=2, value=name)
        ws2.cell(row=row2, column=3, value='Yes' if ws_url else 'No')
        ws2.cell(row=row2, column=4, value='Yes' if li_url else 'No')
        reason_parts = []
        if v.get('source2'):
            reason_parts.append(v['source2'])
        if domain and not ws_conf >= 70:
            reason_parts.append('Email domain: ' + domain)
        if not reason_parts:
            reason_parts.append('No verifiable web presence found')
        ws2.cell(row=row2, column=5, value='; '.join(reason_parts))
        ws2.cell(row=row2, column=6, value=today)
        for col in range(1, 7):
            ws2.cell(row=row2, column=col).border = thin_border
        row2 += 1

# Adjust column widths
for ws in [ws1, ws2]:
    for col_cells in ws.columns:
        max_len = 0
        col_letter = openpyxl.utils.get_column_letter(col_cells[0].column)
        for cell in col_cells:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except:
                pass
        adjusted = min(max_len + 2, 80)
        ws.column_dimensions[col_letter].width = max(adjusted, 15)

wb.save(r'C:\Users\raghe\OneDrive - RATL Technology Ltd\Muhide\CRM_Enriched_Final.xlsx')

enriched_count = row1 - 2
failed_count = row2 - 2
with_ws = sum(1 for r in ws1.iter_rows(min_row=2, values_only=True) if r[8] == 'Yes')
with_li = sum(1 for r in ws1.iter_rows(min_row=2, values_only=True) if r[9] == 'Yes')

print('CRM Enrichment Complete')
print('=' * 50)
print(f'ENRICHED_COMPANIES:  {enriched_count:4d}')
print(f'  With website:      {with_ws:4d}')
print(f'  With LinkedIn:      {with_li:4d}')
print(f'FAILED_LOOKUPS:      {failed_count:4d}')
print(f'Total:               {enriched_count + failed_count:4d}')
print()
print('Manually Verified Companies:')
print(f'  nourmassah.com - Nour Massah Signage (100% confidence)')
print(f'  rafid.com.sa - Rafid Logistics (100% confidence)')
print(f'  site.mohrkeyapp.com - Mohrkey Car Services (100% confidence)')
print(f'  7eleventrading.org - Seven Eleven Trading (100% confidence)')
print(f'  LinkedIn: Mohrkey (100%), Seven Eleven Trading (90%), Nour Massah (85%)')
print(f'  LinkedIn: Bahasan Trading (70%), Mahmas Trading (50%)')
