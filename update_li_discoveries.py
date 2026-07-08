"""Update Sales Intelligence output with LinkedIn discoveries"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

wb = openpyxl.load_workbook('Sales_Intelligence_Output.xlsx')

# New LinkedIn discoveries from web research
new_linkedin = {
    'COMP-000133': {
        'li_url': 'https://sa.linkedin.com/company/megaminditsolutions',
        'li_conf': 100,
        'li_followers': 'Unknown',  # LinkedIn didn't show follower count directly
        'li_employees': '51-100',
        'li_industry': 'IT Services & Healthcare IT',
        'li_description': 'Saudi technology provider: Digital Innovation, Health-tech, Cybersecurity, Infrastructure',
        'li_founded': 2016,
        'ceo': 'Hamza Batterjee',
        'ceo_role': 'President & CEO',
        'employees': '51-100',
        'revenue': '~SAR 66M',
        'headquarters': 'Jeddah, Saudi Arabia',
        'industry': 'IT Services',
        'market_category': 'Technology Supplier',
        'market_group': 'Suppliers',
    },
    'COMP-000718': {
        'li_url': 'https://www.linkedin.com/company/lasetech-ksa',
        'li_conf': 100,
        'li_followers': 297,
        'li_employees': 14,
        'li_industry': 'Medical Equipment Manufacturing',
        'li_description': 'Official distributor of premium American laser and aesthetic devices in Saudi Arabia since 2009',
        'li_founded': 2009,
        'ceo': 'Islam Mostafa',
        'ceo_role': 'Founder & Co-owner',
        'employees': '14',
        'headquarters': 'Jeddah, Saudi Arabia',
        'industry': 'Medical Equipment',
        'market_category': 'Medical Devices',
        'market_group': 'SFDA Regulated',
    },
    'COMP-000069': {
        'li_url': 'https://www.linkedin.com/company/altaawin-medical-group',
        'li_conf': 85,
        'li_followers': 1128,
        'li_employees': 'Unknown',
        'li_industry': 'Medical Services',
        'li_description': 'Healthcare provider with 40+ years of experience in Saudi Arabia',
        'li_founded': 'Unknown',
        'ceo': 'Unknown',
        'employees': 'Unknown',
        'headquarters': 'Saudi Arabia',
        'industry': 'Medical Services',
        'market_category': 'Healthcare Services',
        'market_group': 'SFDA Regulated',
    },
    'COMP-000084': {
        'li_url': 'https://sa.linkedin.com/company/ratl-technology',
        'li_conf': 70,
        'li_followers': 'Unknown',
        'li_employees': 'Unknown',
        'li_industry': 'Construction / Technology',
        'li_description': 'Technology company based in Jeddah, Saudi Arabia',
        'li_founded': 'Unknown',
        'ceo': 'Unknown',
        'ceo_role': '',
        'employees': 'Unknown',
        'headquarters': 'Jeddah, Saudi Arabia',
        'industry': 'Technology',
        'market_category': 'Technology Supplier',
        'market_group': 'Suppliers',
    },
}

# Update SALES_INTELLIGENCE_MASTER
ws1 = wb['SALES_INTELLIGENCE_MASTER']
# Column mapping
SM = {}
for c in range(1, ws1.max_column + 1):
    SM[ws1.cell(1, c).value] = c

updates = 0
for r in range(2, ws1.max_row + 1):
    mid = str(ws1.cell(r, SM['Company Master ID']).value or '')
    if mid in new_linkedin:
        li = new_linkedin[mid]
        ws1.cell(r, SM['LinkedIn URL']).value = li['li_url']
        ws1.cell(r, SM['LinkedIn Confidence']).value = li['li_conf']
        ws1.cell(r, SM['Industry']).value = li['li_industry']
        ws1.cell(r, SM['Market Category']).value = li['market_category']
        ws1.cell(r, SM['Market Group']).value = li['market_group']
        updates += 1

print(f'Updated {updates} records in SALES_INTELLIGENCE_MASTER')

# Update LINKEDIN_DISCOVERY_RESULTS
ws4 = wb['LINKEDIN_DISCOVERY_RESULTS']
# Rebuild this sheet
from openpyxl.utils import get_column_letter

# Clear existing data
ws4.delete_rows(2, ws4.max_row)

hdr = ['Company Master ID', 'Company Name', 'LinkedIn URL', 'LinkedIn Confidence',
       'Followers', 'Employees', 'Industry', 'Description', 'Founded',
       'CEO/Founder', 'CEO Role', 'Decision Maker Email', 'Website Quality',
       'ICP Score', 'TAM Tier', 'Next Actions']

for c, h in enumerate(hdr, 1):
    ws4.cell(1, c).value = h
    ws4.cell(1, c).font = Font(bold=True, color="FFFFFF", size=10)
    ws4.cell(1, c).fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    ws4.cell(1, c).border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                    top=Side(style='thin'), bottom=Side(style='thin'))

def find_company(mid):
    for r in range(2, ws1.max_row + 1):
        if str(ws1.cell(r, SM['Company Master ID']).value or '') == mid:
            return r
    return None

row = 2
for mid, li in sorted(new_linkedin.items()):
    cr = find_company(mid)
    name = ws1.cell(cr, SM['Company Name']).value if cr else mid
    icp = ws1.cell(cr, SM['ICP Score']).value if cr else 0
    tam = ws1.cell(cr, SM['ICP TAM Tier']).value if cr else ''
    wq = ws1.cell(cr, SM['Website Quality']).value if cr else ''
    
    vals = [mid, name, li['li_url'], li['li_conf'],
            li.get('li_followers', ''), li.get('li_employees', ''),
            li['li_industry'], li.get('li_description', '')[:100],
            li.get('li_founded', ''),
            li.get('ceo', ''), li.get('ceo_role', ''),
            '', wq, icp, tam,
            'LinkedIn page confirmed - ready for outreach' if li['li_conf'] >= 100 else 'Verify LinkedIn page manually']
    
    for c, v in enumerate(vals, 1):
        cell = ws4.cell(row, c)
        cell.value = v
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                              top=Side(style='thin'), bottom=Side(style='thin'))
        if c == 4 and v and v >= 100:
            cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        elif c == 4 and v and v >= 85:
            cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    row += 1

# Auto-width
for col_cells in ws4.columns:
    max_len = 0
    col_letter = get_column_letter(col_cells[0].column)
    for cell in col_cells:
        try:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        except:
            pass
    ws4.column_dimensions[col_letter].width = min(max(max_len + 2, 12), 60)

print(f'LINKEDIN_DISCOVERY_RESULTS: {row - 2} records')

# Save
wb.save('Sales_Intelligence_Output.xlsx')
print('Saved to Sales_Intelligence_Output.xlsx')
