import openpyxl, json, sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

# ====== PHASE 0: Understand all source data ======

src = openpyxl.load_workbook('CRM_Enrichment_Output.xlsx')
ws = src['ENRICHMENT_READY']
print(f"ENRICHMENT_READY: {ws.max_row - 1} data rows, {ws.max_column} cols")
headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
for i, h in enumerate(headers, 1):
    print(f"  Col {i}: {h}")

# Correlation: map cols by name
col_map = {}
for c in range(1, ws.max_column + 1):
    col_map[ws.cell(1, c).value] = c
print(f"\nMapping complete: {len(col_map)} columns")

# Phase 1: Scan for corruption
print("\n=== PHASE 1: DATA QUALITY RECOVERY ===")
corruption_log = []
issues = {
    'arabic_in_email': 0,
    'phone_in_name': 0,
    'website_in_name': 0,
    'company_in_contact': 0,
    'invalid_email': 0,
    'missing_email': 0,
    'missing_contact': 0,
    'missing_phone': 0,
    'missing_website': 0,
    'mixed_language_name': 0,
}

import re

def has_arabic(text):
    if not text: return False
    return bool(re.search(r'[\u0600-\u06FF\u0750-\u077F\u08A0-\u08FF\uFB50-\uFDFF\uFE70-\uFEFF]', str(text)))

def count_arabic(text):
    if not text: return 0
    return len(re.findall(r'[\u0600-\u06FF\u0750-\u077F\u08A0-\u08FF\uFB50-\uFDFF\uFE70-\uFEFF]', str(text)))

def count_latin(text):
    if not text: return 0
    return len(re.findall(r'[a-zA-Z]', str(text)))

phone_col = col_map.get('Phone', 23)
email_col = col_map.get('Email', 22)
contact_col = col_map.get('Contact Name', 21)
name_col = col_map.get('Original Company Name', 2)
website_col = col_map.get('Website', 4)
mid_col = col_map.get('Company Master ID', 1)

recovered_contacts = []

for row in range(2, ws.max_row + 1):
    mid = ws.cell(row, mid_col).value or ''
    email = str(ws.cell(row, email_col).value or '')
    phone = str(ws.cell(row, phone_col).value or '')
    contact = str(ws.cell(row, contact_col).value or '')
    name = str(ws.cell(row, name_col).value or '')
    website = str(ws.cell(row, website_col).value or '')

    # Detect Arabic in Email field
    if has_arabic(email) and '@' not in email:
        issues['arabic_in_email'] += 1
        corruption_log.append({
            'row': row, 'mid': mid, 'type': 'Arabic text in Email field',
            'old_value': email, 'detail': 'Contact name stored in Email field'
        })
        recovered_contacts.append({
            'mid': mid,
            'company': name,
            'recovered_contact': email.strip(),
            'original_email': email.strip(),
        })
        # Clear the email field
        ws.cell(row, email_col).value = None

    # Detect phone in company name
    if phone and has_arabic(phone) and len(phone) < 15 and '@' not in phone and count_latin(phone) == 0:
        issues['phone_in_name'] += 1

    # Detect website in company name
    if website and has_arabic(website) and website.startswith('http'):
        pass  # valid website

    # Mixed language in company name
    if has_arabic(name) and count_latin(name) > 3:
        issues['mixed_language_name'] += 1

    # Missing fields
    if not email or email == 'None':
        issues['missing_email'] += 1
    if not contact or contact == 'None':
        issues['missing_contact'] += 1
    if not phone or phone == 'None':
        issues['missing_phone'] += 1
    if not website or website == 'None':
        issues['missing_website'] += 1

    # Invalid email format (has @ but wrong)
    if email and '@' in email:
        parts = email.split('@')
        if len(parts) != 2 or not parts[1] or '.' not in parts[1]:
            issues['invalid_email'] += 1

print("\nCorruption Detection Results:")
for k, v in sorted(issues.items()):
    print(f"  {k}: {v}")

print(f"\nRecovered Contacts: {len(recovered_contacts)}")
for rc in recovered_contacts[:10]:
    print(f"  {rc['mid']}: '{rc['original_email']}' -> Contact Name")

# Save recovery log
json.dump(recovered_contacts, open('recovered_contacts.json', 'w', encoding='utf-8'), ensure_ascii=False, indent=2)

# Save cleaned source
src.save('CRM_Enrichment_Output_CLEANED.xlsx')
print("\nCleaned source saved to CRM_Enrichment_Output_CLEANED.xlsx")
