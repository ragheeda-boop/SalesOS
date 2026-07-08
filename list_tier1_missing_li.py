import openpyxl

wb = openpyxl.load_workbook('Sales_Intelligence_Output.xlsx')
ws_t1 = wb['ICP_PRIORITY_ACCOUNTS']

t1_accounts = []
for r in range(2, ws_t1.max_row + 1):
    t1_accounts.append({
        'id': str(ws_t1.cell(r, 1).value or ''),
        'name': str(ws_t1.cell(r, 2).value or ''),
        'arabic_name': str(ws_t1.cell(r, 3).value or ''),
        'score': ws_t1.cell(r, 4).value,
    })

wb2 = openpyxl.load_workbook('Website_LinkedIn_Enrichment.xlsx')
ws_li = wb2['LINKEDIN_VERIFIED']
li_found = set()
for r in range(2, ws_li.max_row + 1):
    try:
        li_found.add(str(ws_li.cell(r, 1).value or ''))
    except:
        pass

ws_wv = wb2['WEBSITE_VERIFIED']
website_info = {}
for r in range(2, ws_wv.max_row + 1):
    try:
        mid = str(ws_wv.cell(r, 1).value or '')
        website_info[mid] = {
            'url': str(ws_wv.cell(r, 4).value or ''),
            'title': str(ws_wv.cell(r, 8).value or ''),
            'desc': str(ws_wv.cell(r, 9).value or ''),
        }
    except:
        pass

ws_wnf = wb2['WEBSITE_NOT_FOUND']
for r in range(2, ws_wnf.max_row + 1):
    try:
        mid = str(ws_wnf.cell(r, 1).value or '')
        domain = str(ws_wnf.cell(r, 5).value or '')
        if mid not in website_info and domain:
            website_info[mid] = {'url': 'http://' + domain, 'title': '', 'desc': ''}
    except:
        pass

print("TIER 1 ACCOUNTS - LINKEDIN STATUS")
print("=" * 80)

have_li = []
need_li = []

for a in t1_accounts:
    if a['id'] in li_found:
        have_li.append(a)
    else:
        need_li.append(a)

print(f"\nAlready have LinkedIn (SKIP): {len(have_li)}")
for a in have_li:
    print(f"  {a['id']} | {a['name']}")

print(f"\n\nNEED LINKEDIN: {len(need_li)} accounts")
print("=" * 80)

for i, a in enumerate(need_li, 1):
    wi = website_info.get(a['id'], {})
    ws_url = wi.get('url', '') if wi else ''
    print(f"\n[{i}/{len(need_li)}] {a['id']}")
    print(f"    Company:      {a['name']}")
    print(f"    Arabic:       {a['arabic_name']}")
    print(f"    ICP Score:    {a['score']}")
    print(f"    Website:      {ws_url if ws_url else 'NONE'}")
    print(f"    Search Queries:")
    print(f"      1. site:linkedin.com/company \"{a['name']}\"")
    print(f"      2. \"{a['name']}\" LinkedIn")
    if a['arabic_name'] and a['arabic_name'] != 'None':
        print(f"      3. \"{a['arabic_name']}\" لينكدإن")
    if ws_url:
        domain = ws_url.replace('http://', '').replace('https://', '').split('/')[0]
        print(f"      4. site:{domain} linkedin")

print(f"\n\nSUMMARY:")
print(f"  Total Tier 1: {len(t1_accounts)}")
print(f"  With LinkedIn: {len(have_li)}")
print(f"  Missing LinkedIn: {len(need_li)}")
print(f"  Target (50%): {int(len(t1_accounts) * 0.5)} accounts with LI")
print(f"  Need to find: {int(len(t1_accounts) * 0.5) - len(have_li)} more")
