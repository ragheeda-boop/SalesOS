import openpyxl

wb = openpyxl.load_workbook('CRM_Enrichment_Output.xlsx')
print('Sheets:', wb.sheetnames)

ws = wb['ENRICHMENT_READY']
print(f'Rows: {ws.max_row}, Cols: {ws.max_column}')

print('Headers:')
for col in range(1, ws.max_column+1):
    h = ws.cell(1, col).value
    print(f'  Col {col}: {h}')

print()
print('First 5 data rows (cols 1-10):')
for row in range(2, min(7, ws.max_row+1)):
    vals = [str(ws.cell(row, c).value or '')[:50] for c in range(1, 11)]
    print(f'  Row {row}: {" | ".join(vals)}')

# Get all company names
print()
print(f'Total companies to enrich: {ws.max_row - 1}')
names = []
for row in range(2, ws.max_row+1):
    mid = ws.cell(row, 1).value or ''
    name = ws.cell(row, 3).value or ''
    names.append((mid, name))

print('Sample companies:')
for mid, name in names[:20]:
    print(f'  {mid}: {name}')
