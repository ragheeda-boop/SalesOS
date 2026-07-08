import openpyxl, json, re

wb = openpyxl.load_workbook('Sales_Intelligence_Output_COPY.xlsx')
ws_t1 = wb['ICP_PRIORITY_ACCOUNTS']

wb2 = openpyxl.load_workbook('Website_LinkedIn_Enrichment.xlsx')
ws_li = wb2['LINKEDIN_VERIFIED']
li_found = {}
for r in range(2, ws_li.max_row + 1):
    mid = ws_li.cell(r, 1).value
    if mid:
        li_found[str(mid)] = {
            'url': str(ws_li.cell(r, 4).value or ''),
            'conf': ws_li.cell(r, 5).value or 0
        }

ws_wv = wb2['WEBSITE_VERIFIED']
website_info = {}
for r in range(2, ws_wv.max_row + 1):
    mid = ws_wv.cell(r, 1).value
    if mid:
        website_info[str(mid)] = str(ws_wv.cell(r, 4).value or '')

t1_accounts = []
for r in range(2, ws_t1.max_row + 1):
    mid = str(ws_t1.cell(r, 1).value or '')
    name = str(ws_t1.cell(r, 2).value or '')
    arabic = str(ws_t1.cell(r, 3).value or '')
    score = ws_t1.cell(r, 4).value
    li = li_found.get(mid, {})
    t1_accounts.append({
        'id': mid, 'name': name, 'arabic': arabic, 'score': score,
        'has_li': mid in li_found,
        'li_url': li.get('url', ''),
        'li_conf': li.get('conf', 0),
        'website': website_info.get(mid, '')
    })

with open('tier1_status.txt', 'w', encoding='utf-8') as f:
    f.write("TIER 1 WITH LINKEDIN (SKIP):\n")
    for a in t1_accounts:
        if a['has_li']:
            f.write(f"  {a['id']} | {a['name']} | {a['li_url']} (conf: {a['li_conf']})\n")

    need = [a for a in t1_accounts if not a['has_li']]
    f.write(f"\n\nTIER 1 NEEDING LINKEDIN ({len(need)}):\n")
    f.write("=" * 100 + "\n")
    for i, a in enumerate(need, 1):
        f.write(f"[{i:2d}/{len(need)}] {a['id']} | Score: {a['score']}\n")
        f.write(f"    Company: {a['name']}\n")
        f.write(f"    Arabic:  {a['arabic']}\n")
        if a['website']:
            f.write(f"    Website: {a['website']}\n")
        # Generate search queries
        f.write(f"    S1: site:linkedin.com/company \"{a['name']}\"\n")
        f.write(f"    S2: \"{a['name']}\" LinkedIn\n")
        if a['arabic'] and a['arabic'] != 'None':
            f.write(f"    S3: \"{a['arabic']}\" لينكدإن\n")
        if a['website']:
            from urllib.parse import urlparse
            domain = urlparse(a['website']).netloc or a['website'].replace('http://','').replace('https://','').split('/')[0]
            f.write(f"    S4: site:{domain} linkedin\n")
        f.write("\n")

    f.write(f"\n\nSUMMARY:\n")
    f.write(f"  Total Tier 1: {len(t1_accounts)}\n")
    f.write(f"  With LinkedIn: {len([a for a in t1_accounts if a['has_li']])}\n")
    f.write(f"  Missing LinkedIn: {len(need)}\n")
    f.write(f"  Target (50%): {int(len(t1_accounts) * 0.5)} accounts with LI\n")
    f.write(f"  Need to find: {int(len(t1_accounts) * 0.5) - len([a for a in t1_accounts if a['has_li']])} more\n")

print("Written to tier1_status.txt")
print(f"Total Tier 1: {len(t1_accounts)}")
print(f"Have LI: {len([a for a in t1_accounts if a['has_li']])}")
print(f"Need LI: {len([a for a in t1_accounts if not a['has_li']])}")
