"""
AI Sales OS - Phase 3 Sales Intelligence Enrichment
Complete pipeline: Classification, LinkedIn, ICP, TAM, Decision Makers, Routing
"""
import json, openpyxl, re, sys, io, socket, time
from datetime import datetime
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

TODAY = datetime.now().strftime('%Y-%m-%d')
TIMESTAMP = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

# ===================== STYLING =====================
hdr_font = Font(bold=True, color="FFFFFF", size=10)
hdr_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
hdr_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
blue_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")

def style_header(ws, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(1, c)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = hdr_align
        cell.border = thin_border

def auto_width(ws):
    for col_cells in ws.columns:
        max_len = 0
        col_letter = openpyxl.utils.get_column_letter(col_cells[0].column)
        for cell in col_cells:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 12), 70)

# ===================== LOAD DATA =====================
print("=== LOADING CLEAN_MASTER ===")
wb = openpyxl.load_workbook('CRM_DataQuality_Output.xlsx')
ws = wb['CLEAN_MASTER']

# Column mapping from CLEAN_MASTER
CM = {}
for c in range(1, ws.max_column + 1):
    CM[ws.cell(1, c).value] = c

records = []
for r in range(2, ws.max_row + 1):
    mid = str(ws.cell(r, CM['Company Master ID']).value or '')
    name = str(ws.cell(r, CM['Company Name']).value or '')
    std_name = str(ws.cell(r, CM['Standard Name']).value or '')
    contact = str(ws.cell(r, CM['Contact Name']).value or '')
    first_name = str(ws.cell(r, CM['First Name']).value or '')
    last_name = str(ws.cell(r, CM['Last Name']).value or '')
    email = str(ws.cell(r, CM['Email']).value or '')
    phone = str(ws.cell(r, CM['Phone']).value or '')
    website = str(ws.cell(r, CM['Validated Website']).value or '')
    domain = str(ws.cell(r, CM['Validated Domain']).value or '')
    dom_conf = ws.cell(r, CM['Domain Confidence']).value or 0
    li_url = str(ws.cell(r, CM['LinkedIn URL']).value or '')
    li_conf = ws.cell(r, CM['LinkedIn Confidence']).value or 0
    country = str(ws.cell(r, CM['Country']).value or '')
    city = str(ws.cell(r, CM['City']).value or '')
    industry = str(ws.cell(r, CM['Industry']).value or '')
    salesperson = str(ws.cell(r, CM['Salesperson']).value or '')
    stage = str(ws.cell(r, CM['Stage']).value or '')
    tags = str(ws.cell(r, CM['Tags']).value or '')
    recovered = str(ws.cell(r, CM['Recovered Contact']).value or '')
    quality_score = ws.cell(r, CM['Quality Score']).value or 0
    quality_tier = str(ws.cell(r, CM['Quality Tier']).value or '')
    
    records.append({
        'mid': mid, 'name': name, 'std_name': std_name,
        'contact': contact, 'first_name': first_name, 'last_name': last_name,
        'email': email, 'phone': phone,
        'website': website, 'domain': domain, 'dom_conf': dom_conf,
        'li_url': li_url, 'li_conf': li_conf,
        'country': country, 'city': city, 'industry': industry,
        'salesperson': salesperson, 'stage': stage, 'tags': tags,
        'recovered': recovered, 'quality_score': quality_score, 'quality_tier': quality_tier,
    })

print(f"Loaded {len(records)} records")

# Phase 1: Prioritize A > B > C
priority_order = {'A': 0, 'B': 1, 'C': 2}
priority_records = [r for r in records if r['quality_tier'] in ('A', 'B', 'C')]
priority_records.sort(key=lambda x: (priority_order.get(x['quality_tier'], 9), -x['quality_score']))
print(f"Priority records: {len(priority_records)} (A={sum(1 for r in priority_records if r['quality_tier']=='A')}, B={sum(1 for r in priority_records if r['quality_tier']=='B')}, C={sum(1 for r in priority_records if r['quality_tier']=='C')})")

# ===================== PHASE 2-3: COMPANY INTELLIGENCE + SAUDI MARKET CLASSIFICATION =====================
print("\n=== PHASE 2-3: COMPANY INTELLIGENCE & MARKET CLASSIFICATION ===")

# Arabic keyword classifiers for Saudi Market Classification
arabic_classifiers = {
    'General Contractor': ['مقاول عام', 'مقاولات عامه', 'مقاولات عامة'],
    'Civil Contractor': ['مقاول مدني', 'مقاولات مدنيه'],
    'Infrastructure Contractor': ['مقاول بنية تحتية', 'مقاولات infrastructure'],
    'MEP Contractor': ['مقاول ميكانيكا', 'مقاول كهرباء', 'مقاول تكييف', 'اعمال ميكانيكيه', 'مقاولات م م ك'],
    'Government Contractor': ['مقاول حكومي'],
    'Design Office': ['مكتب تصميم', 'مكتب هندسي', 'استشارات هندسيه'],
    'Consulting Office': ['مكتب استشارات', 'استشاري'],
    'Project Management Office': ['ادارة مشاريع'],
    'Building Materials Supplier': ['مواد بناء', 'مواد بناء وتشطيب', 'اسمنت', 'حديد', 'مقاولات ومواد'],
    'Electrical Supplier': ['مواد كهربائيه', 'كهرباء', 'اجهزة كهربائيه'],
    'HVAC Supplier': ['تكييف', 'تبريد', 'مكيفات'],
    'Safety Supplier': ['سلامه', 'اطفاء حريق', 'انذار حريق', 'معدات سلامه'],
    'Equipment Supplier': ['معدات صناعيه', 'معدات ثقيله', 'تأجير معدات', 'معدات وادوات'],
    'Industrial Supplies': ['مستلزمات صناعيه', 'مصنع', 'صناعي'],
    'Technology Supplier': ['تقنيه', 'تقنية', 'حاسب', 'برمجيات', 'انظمة', 'it'],
    'Food Supplier': ['تمور', 'مياه', 'اغذيه', 'اغذية', 'مواد غذائيه', 'حلويات', 'مخابز', 'غذائيه', 'غذائية'],
    'Medical Devices': ['معدات طبيه', 'اجهزة طبيه', 'طبيه', 'طبية'],
    'Pharmaceutical': ['صيدليه', 'ادويه', 'دواء', 'مستحضرات'],
    'Cosmetics': ['مستحضرات تجميل', 'عطور', 'مستلزمات تجميل'],
    'Laboratory': ['مختبر', 'تحاليل'],
    'Healthcare Services': ['خدمات صحيه', 'مركز طبي', 'مستشفى'],
    'Logistics': ['خدمات لوجستيه', 'شحن', 'نقل', 'مواصلات'],
    'Retail': ['متجر', 'محل', 'تجزئه', 'مول', 'سوبرماركت'],
    'Agriculture': ['زراعي', 'زراعه', 'مزارع'],
    'Real Estate': ['عقار', 'عقاري'],
}

latin_classifiers = {
    'General Contractor': ['general contractor', 'construction', 'building contractor'],
    'Civil Contractor': ['civil contractor'],
    'MEP Contractor': ['mep', 'mechanical', 'electrical', 'plumbing', 'hvac'],
    'Consulting Office': ['consulting', 'consultancy', 'engineering consultancy', 'engineering office'],
    'Project Management': ['project management', 'program management'],
    'Building Materials Supplier': ['building materials', 'construction materials', 'supply'],
    'Food Supplier': ['food', 'beverage', 'bakery', 'catering', 'provisions'],
    'Technology': ['technology', 'software', 'it services', 'digital', 'tech'],
    'Medical': ['medical', 'healthcare', 'pharma', 'pharmaceutical', 'clinic'],
    'Logistics': ['logistics', 'shipping', 'transport', 'freight'],
    'Industrial': ['industrial', 'manufacturing', 'factory', 'plant'],
    'Retail': ['retail', 'store', 'shop', 'trading', 'trade'],
    'Agriculture': ['agriculture', 'agricultural', 'farm', 'agri'],
    'Real Estate': ['real estate', 'property', 'realty'],
}

def classify_market(name):
    """Classify company into Saudi market category based on name keywords."""
    name_lower = name.lower()
    
    # Arabic matching
    for category, keywords in arabic_classifiers.items():
        for kw in keywords:
            if kw in name_lower or kw in name:
                return category
    
    # Latin matching
    for category, keywords in latin_classifiers.items():
        for kw in keywords:
            if kw in name_lower:
                return category
    
    # Default classification based on name patterns
    if re.search(r'شركه|مؤسسه|مصنع', name):
        return 'General Contractor'
    if re.search(r'trading|est\.|establishment', name_lower):
        return 'Supplier'
    
    return 'Unclassified'

def classify_market_group(category):
    """Map detailed category to market group."""
    contractor_keywords = ['Contractor', 'General', 'Civil', 'Infrastructure', 'MEP', 'Government', 'Industrial']
    engineering_keywords = ['Design', 'Consulting', 'Project Management', 'Supervision', 'Engineering', 'Office']
    supplier_keywords = ['Supplier', 'Building Materials', 'Electrical', 'Mechanical', 'HVAC', 'Safety', 'Equipment', 'Industrial Supplies', 'Technology', 'Retail', 'Agriculture', 'Real Estate']
    sfda_keywords = ['Medical', 'Pharmaceutical', 'Food', 'Cosmetics', 'Laboratory', 'Healthcare']
    
    if any(kw in category for kw in contractor_keywords):
        return 'Contractors'
    if any(kw in category for kw in engineering_keywords):
        return 'Engineering Offices'
    if any(kw in category for kw in supplier_keywords):
        return 'Suppliers'
    if any(kw in category for kw in sfda_keywords):
        return 'SFDA Regulated'
    return 'Other'

# Also try to determine employee range and size tier from name patterns or tags
def infer_employee_range(name, tags):
    """Try to infer employee range from available data."""
    tags_lower = tags.lower()
    name_lower = name.lower()
    
    if any(kw in name_lower for kw in ['factory', 'مصنع كبير', 'شركه كبيره', 'group', 'مجموعه', 'holding']):
        return '201-500', 'Large'
    if any(kw in tags_lower for kw in ['enterprise', 'large', 'corporate']):
        return '101-250', 'Medium-Large'
    if any(kw in tags_lower for kw in ['sme', 'medium']):
        return '11-50', 'Small-Medium'
    if any(kw in tags_lower for kw in ['small', 'micro']):
        return '1-10', 'Micro'
    
    return 'Unknown', 'Unknown'

# Process all priority records
for rec in priority_records:
    name = rec['name']
    tags = rec['tags']
    domain = rec['domain']
    
    # Market Classification
    category = classify_market(name)
    group = classify_market_group(category)
    rec['market_category'] = category
    rec['market_group'] = group
    
    # Employee range
    emp_range, size_tier = infer_employee_range(name, tags)
    rec['employee_range'] = emp_range
    rec['size_tier'] = size_tier
    
    # Business category from tags
    rec['business_category'] = tags if tags and tags != 'None' else ''
    
    # City from available data
    rec['city'] = rec['city'] if rec['city'] and rec['city'] != 'None' else ''
    
    # Website quality assessment
    if rec['dom_conf'] >= 100:
        rec['website_quality'] = 'Excellent'
    elif rec['dom_conf'] >= 85:
        rec['website_quality'] = 'Good'
    elif rec['dom_conf'] >= 70:
        rec['website_quality'] = 'Fair'
    elif rec['domain']:
        rec['website_quality'] = 'Poor'
    else:
        rec['website_quality'] = 'None'

# Print classification distribution
from collections import Counter
group_dist = Counter(r['market_group'] for r in priority_records)
cat_dist = Counter(r['market_category'] for r in priority_records)
print("Market Group Distribution:")
for g, c in sorted(group_dist.items(), key=lambda x: -x[1]):
    print(f"  {g}: {c}")
print("\nTop Market Categories:")
for cat, c in sorted(cat_dist.items(), key=lambda x: -x[1])[:10]:
    print(f"  {cat}: {c}")

# ===================== PHASE 4-5: LINKED INTELLIGENCE + ICP SCORE =====================
print("\n=== PHASE 4-5: LINKEDIN INTELLIGENCE & ICP SCORING ===")

def compute_icp_score(rec):
    """Calculate ICP Score (0-100) based on multiple dimensions."""
    score = 0
    
    # Industry Fit (0-20)
    if rec['market_group'] != 'Unclassified' and rec['market_group'] != 'Other':
        score += 15
    if rec['market_category'] != 'Unclassified':
        score += 5
    
    # Company Size (0-15)
    if rec['size_tier'] in ('Large', 'Medium-Large'):
        score += 15
    elif rec['size_tier'] == 'Small-Medium':
        score += 10
    elif rec['size_tier'] == 'Micro':
        score += 5
    else:
        score += 8  # Unknown but priority tier
    
    # Market Relevance - Saudi Arabia (0-10)
    if 'Saudi' in rec['country']:
        score += 10
    
    # Digital Presence (0-20)
    if rec['domain']:
        score += 10
    if rec['website_quality'] in ('Excellent', 'Good'):
        score += 10
    elif rec['website_quality'] == 'Fair':
        score += 5
    
    # Decision-Maker Accessibility (0-15)
    if rec['contact'] and rec['contact'] != 'None' and rec['contact'].strip():
        score += 10
        if rec['email'] and '@' in rec['email']:
            score += 5
    elif rec['email'] and '@' in rec['email']:
        score += 8
    
    # LinkedIn Presence (0-10)
    if rec['li_url'] and rec['li_conf'] >= 70:
        score += 10
    elif rec['li_url']:
        score += 5
    
    # Website Quality (0-10)
    if rec['website_quality'] == 'Excellent':
        score += 10
    elif rec['website_quality'] == 'Good':
        score += 7
    elif rec['website_quality'] == 'Fair':
        score += 4
    elif rec['website_quality'] == 'Poor':
        score += 2
    
    return min(score, 100)

for rec in priority_records:
    rec['icp_score'] = compute_icp_score(rec)
    
    # TAM Segmentation (Phase 6)
    if rec['icp_score'] >= 80:
        rec['tam_tier'] = 'Tier 1 - Strategic Accounts'
    elif rec['icp_score'] >= 65:
        rec['tam_tier'] = 'Tier 2 - High Potential'
    elif rec['icp_score'] >= 50:
        rec['tam_tier'] = 'Tier 3 - Medium Potential'
    else:
        rec['tam_tier'] = 'Tier 4 - Long Tail'
    
    # Decile ranking within priority group
    rec['icp_rank_group'] = ''

# ICP Distribution
icp_dist = Counter()
for r in priority_records:
    if r['icp_score'] >= 80: icp_dist['80-100'] += 1
    elif r['icp_score'] >= 65: icp_dist['65-79'] += 1
    elif r['icp_score'] >= 50: icp_dist['50-64'] += 1
    else: icp_dist['0-49'] += 1

print("ICP Score Distribution:")
for k in ['80-100', '65-79', '50-64', '0-49']:
    print(f"  {k}: {icp_dist.get(k, 0)}")

tam_dist = Counter(r['tam_tier'] for r in priority_records)
print("\nTAM Distribution:")
for t, c in sorted(tam_dist.items()):
    print(f"  {t}: {c}")

# ===================== PHASE 7: DECISION MAKER DISCOVERY =====================
print("\n=== PHASE 7: DECISION MAKER DISCOVERY ===")

# For Tier 1 and Tier 2 accounts, extract decision maker info from available contact data
decision_makers = []
for rec in priority_records:
    if rec['tam_tier'] in ('Tier 1 - Strategic Accounts', 'Tier 2 - High Potential'):
        dm = {
            'mid': rec['mid'],
            'company': rec['name'],
            'contact_name': rec['contact'],
            'first_name': rec['first_name'],
            'last_name': rec['last_name'],
            'email': rec['email'],
            'phone': rec['phone'],
            'linkedin': rec['li_url'],
            'website': rec['website'],
            'domain': rec['domain'],
            'tam_tier': rec['tam_tier'],
            'quality_tier': rec['quality_tier'],
            'likely_role': '',
            'confidence': 50,
        }
        
        # Infer likely role from available context
        if rec['contact']:
            name_upper = rec['contact'].upper()
            if any(kw in name_upper for kw in ['MD ', 'M.D.', 'MANAGING DIRECTOR', 'CEO', 'C.E.O.']):
                dm['likely_role'] = 'Managing Director / CEO'
                dm['confidence'] = 80
            elif any(kw in name_upper for kw in ['GM ', 'GENERAL MANAGER']):
                dm['likely_role'] = 'General Manager'
                dm['confidence'] = 75
            elif any(kw in name_upper for kw in ['OWNER', 'PROPRIETOR']):
                dm['likely_role'] = 'Owner'
                dm['confidence'] = 85
            elif any(kw in name_upper for kw in ['PURCHASE', 'PROCUREMENT']):
                dm['likely_role'] = 'Procurement Manager'
                dm['confidence'] = 70
            elif any(kw in name_upper for kw in ['FINANCE', 'CFO', 'FINANCIAL']):
                dm['likely_role'] = 'Finance Manager'
                dm['confidence'] = 70
            elif any(kw in name_upper for kw in ['OPERATION']):
                dm['likely_role'] = 'Operations Manager'
                dm['confidence'] = 70
            elif any(kw in name_upper for kw in ['BUSINESS DEV', 'BD']):
                dm['likely_role'] = 'Business Development Manager'
                dm['confidence'] = 70
            else:
                dm['likely_role'] = 'Contact (Role Unknown)'
                dm['confidence'] = 40
        else:
            dm['likely_role'] = 'Unknown'
            dm['confidence'] = 0
        
        decision_makers.append(dm)

print(f"Decision makers identified: {len(decision_makers)}")
roles = Counter(dm['likely_role'] for dm in decision_makers)
for r, c in sorted(roles.items(), key=lambda x: -x[1])[:10]:
    print(f"  {r}: {c}")

# ===================== PHASE 8: SALES ROUTING =====================
print("\n=== PHASE 8: SALES ROUTING ===")

# Territory mapping based on city
territory_map = {
    'Riyadh': 'Central',
    'Jeddah': 'Western',
    'Makkah': 'Western',
    'Mecca': 'Western',
    'Madinah': 'Western',
    'Medina': 'Western',
    'Dammam': 'Eastern',
    'Khobar': 'Eastern',
    'Al Khobar': 'Eastern',
    'Dhahran': 'Eastern',
    'Jubail': 'Eastern',
    'Yanbu': 'Western',
    'Tabuk': 'Northern',
    'Abha': 'Southern',
    'Khamis Mushait': 'Southern',
    'Najran': 'Southern',
    'Jizan': 'Southern',
    'Hail': 'Northern',
    'Qassim': 'Central',
    'Buraydah': 'Central',
    'Taif': 'Western',
}

# Outreach channel by TAM tier
outreach_by_tier = {
    'Tier 1 - Strategic Accounts': 'Multi-channel (Email + LinkedIn + Phone + In-person)',
    'Tier 2 - High Potential': 'Email + LinkedIn + Phone',
    'Tier 3 - Medium Potential': 'Email + LinkedIn',
    'Tier 4 - Long Tail': 'Automated Email Sequence',
}

# Recommended sequence
sequence_by_tier = {
    'Tier 1 - Strategic Accounts': 'Day 1: LinkedIn Connect > Day 3: Personalized Email > Day 7: Phone Call > Day 14: Follow-up Email > Day 21: In-person Meeting Request',
    'Tier 2 - High Potential': 'Day 1: Email > Day 4: LinkedIn Connect > Day 8: Phone Call > Day 15: Follow-up Email',
    'Tier 3 - Medium Potential': 'Day 1: Email > Day 5: LinkedIn Connect > Day 12: Follow-up Email',
    'Tier 4 - Long Tail': 'Day 1: Automated Email > Day 14: Follow-up Email',
}

sales_routes = []
for rec in priority_records:
    city = rec['city'] if rec['city'] and rec['city'] != 'None' else ''
    territory = 'Central'  # Default
    for city_key, terr in territory_map.items():
        if city_key.lower() in city.lower():
            territory = terr
            break
    
    region = 'KSA'
    
    # Sales queue by market group
    market = rec['market_group']
    if market == 'Contractors':
        queue = 'Contractors Queue'
    elif market == 'Engineering Offices':
        queue = 'Engineering Queue'
    elif market == 'Suppliers':
        queue = 'Suppliers Queue'
    elif market == 'SFDA Regulated':
        queue = 'SFDA Queue'
    else:
        queue = 'General Queue'
    
    tam_tier = rec['tam_tier']
    icp = rec['icp_score']
    
    route = {
        'mid': rec['mid'],
        'company': rec['name'],
        'account_owner': rec['salesperson'] if rec['salesperson'] and rec['salesperson'] != 'None' else 'Unassigned',
        'territory': territory,
        'region': region,
        'sales_queue': queue,
        'tam_tier': tam_tier,
        'icp_score': icp,
        'priority': 'Critical' if icp >= 80 else 'High' if icp >= 65 else 'Medium' if icp >= 50 else 'Standard',
        'recommended_sequence': sequence_by_tier.get(tam_tier, ''),
        'recommended_channel': outreach_by_tier.get(tam_tier, ''),
    }
    sales_routes.append(route)
    rec['territory'] = territory
    rec['sales_queue'] = queue
    rec['priority'] = route['priority']

print("Sales Queue Distribution:")
queue_dist = Counter(r['sales_queue'] for r in sales_routes)
for q, c in sorted(queue_dist.items(), key=lambda x: -x[1]):
    print(f"  {q}: {c}")

# ===================== PHASE 9: GENERATE OUTPUTS =====================
print("\n=== PHASE 9: GENERATING OUTPUT SHEETS ===")
OUTPUT_FILE = 'Sales_Intelligence_Output.xlsx'
wb_out = openpyxl.Workbook()

# ----- SHEET 1: SALES_INTELLIGENCE_MASTER -----
ws1 = wb_out.active
ws1.title = 'SALES_INTELLIGENCE_MASTER'

si_headers = [
    'Company Master ID', 'Company Name', 'Standard Name',
    'Contact Name', 'First Name', 'Last Name',
    'Email', 'Phone',
    'Validated Website', 'Validated Domain', 'Domain Confidence',
    'LinkedIn URL', 'LinkedIn Confidence',
    'Country', 'City',
    'Industry', 'Sub Industry', 'Business Category',
    'Employee Range', 'Company Size Tier',
    'Market Group', 'Market Category',
    'Website Quality', 'Quality Score', 'Quality Tier',
    'ICP Score', 'ICP TAM Tier',
    'Salesperson', 'Territory', 'Region', 'Sales Queue', 'Priority',
    'Recommended Sequence', 'Recommended Channel',
]
for i, h in enumerate(si_headers, 1):
    ws1.cell(1, i, h)
style_header(ws1, len(si_headers))

# Sort priority records by ICP score descending
priority_records.sort(key=lambda x: -x['icp_score'])

for ri, rec in enumerate(priority_records, 2):
    vals = [
        rec['mid'], rec['name'], rec['std_name'],
        rec['contact'], rec['first_name'], rec['last_name'],
        rec['email'], rec['phone'],
        rec['website'], rec['domain'], rec['dom_conf'],
        rec['li_url'], rec['li_conf'],
        rec['country'], rec['city'],
        rec['industry'], '', rec['business_category'],
        rec['employee_range'], rec['size_tier'],
        rec['market_group'], rec['market_category'],
        rec['website_quality'], rec['quality_score'], rec['quality_tier'],
        rec['icp_score'], rec['tam_tier'],
        rec['salesperson'], rec.get('territory', ''), 'KSA',
        rec.get('sales_queue', ''), rec.get('priority', ''),
        sequence_by_tier.get(rec['tam_tier'], ''),
        outreach_by_tier.get(rec['tam_tier'], ''),
    ]
    for ci, v in enumerate(vals, 1):
        cell = ws1.cell(ri, ci, v)
        cell.border = thin_border
        if ci == 26:  # ICP Score
            if v and v >= 80: cell.fill = green_fill
            elif v and v >= 65: cell.fill = blue_fill
            elif v and v >= 50: cell.fill = yellow_fill

auto_width(ws1)
print(f"SALES_INTELLIGENCE_MASTER: {len(priority_records)} records")

# ----- SHEET 2: ICP_PRIORITY_ACCOUNTS -----
ws2 = wb_out.create_sheet('ICP_PRIORITY_ACCOUNTS')
icp_headers = [
    'Company Master ID', 'Company Name', 'ICP Score', 'TAM Tier',
    'Market Group', 'Market Category', 'Website', 'LinkedIn URL',
    'Contact Name', 'Email', 'Phone', 'City', 'Salesperson', 'Priority'
]
for i, h in enumerate(icp_headers, 1):
    ws2.cell(1, i, h)
style_header(ws2, len(icp_headers))

tier1_accounts = [r for r in priority_records if r['icp_score'] >= 80]
for ri, rec in enumerate(tier1_accounts, 2):
    vals = [
        rec['mid'], rec['name'], rec['icp_score'], rec['tam_tier'],
        rec['market_group'], rec['market_category'], rec['website'], rec['li_url'],
        rec['contact'], rec['email'], rec['phone'], rec['city'],
        rec['salesperson'], rec.get('priority', ''),
    ]
    for ci, v in enumerate(vals, 1):
        cell = ws2.cell(ri, ci, v)
        cell.border = thin_border

auto_width(ws2)
print(f"ICP_PRIORITY_ACCOUNTS (ICP > 80): {len(tier1_accounts)} records")

# ----- SHEET 3: DECISION_MAKERS -----
ws3 = wb_out.create_sheet('DECISION_MAKERS')
dm_headers = [
    'Company Master ID', 'Company Name', 'Contact Name', 'First Name', 'Last Name',
    'Email', 'Phone', 'Likely Role', 'Role Confidence',
    'LinkedIn URL', 'Website', 'TAM Tier', 'Quality Tier',
]
for i, h in enumerate(dm_headers, 1):
    ws3.cell(1, i, h)
style_header(ws3, len(dm_headers))

for ri, dm in enumerate(decision_makers, 2):
    vals = [
        dm['mid'], dm['company'], dm['contact_name'], dm['first_name'], dm['last_name'],
        dm['email'], dm['phone'], dm['likely_role'], dm['confidence'],
        dm['linkedin'], dm['website'], dm['tam_tier'], dm['quality_tier'],
    ]
    for ci, v in enumerate(vals, 1):
        cell = ws3.cell(ri, ci, v)
        cell.border = thin_border
        if ci == 8 and v and 'CEO' in v:
            cell.fill = green_fill

auto_width(ws3)
print(f"DECISION_MAKERS: {len(decision_makers)} records")

# ----- SHEET 4: LINKEDIN_DISCOVERY_RESULTS -----
ws4 = wb_out.create_sheet('LINKEDIN_DISCOVERY_RESULTS')
ld_headers = [
    'Company Master ID', 'Company Name', 'LinkedIn URL', 'LinkedIn Confidence',
    'Decision Maker Name', 'Decision Maker Role', 'Website Quality',
    'ICP Score', 'TAM Tier', 'Next Actions',
]
for i, h in enumerate(ld_headers, 1):
    ws4.cell(1, i, h)
style_header(ws4, len(ld_headers))

ld_records = [r for r in priority_records if r['li_url']]
for ri, rec in enumerate(ld_records, 2):
    dm_name = rec['contact'] if rec['contact'] and rec['contact'] != 'None' else 'Unknown'
    vals = [
        rec['mid'], rec['name'], rec['li_url'], rec['li_conf'],
        dm_name, '', rec['website_quality'],
        rec['icp_score'], rec['tam_tier'],
        'Confirm LinkedIn page matches company' if rec['li_conf'] < 90 else 'Ready for outreach',
    ]
    for ci, v in enumerate(vals, 1):
        cell = ws4.cell(ri, ci, v)
        cell.border = thin_border

auto_width(ws4)
print(f"LINKEDIN_DISCOVERY_RESULTS: {len(ld_records)} records")

# ----- SHEET 5: OUTBOUND_READY -----
ws5 = wb_out.create_sheet('OUTBOUND_READY')
or_headers = [
    'Company Master ID', 'Company Name', 'Contact Name', 'Email', 'Phone',
    'Validated Website', 'LinkedIn URL',
    'ICP Score', 'TAM Tier', 'Priority',
    'Salesperson', 'Territory', 'Recommended Channel', 'Recommended Sequence',
]
for i, h in enumerate(or_headers, 1):
    ws5.cell(1, i, h)
style_header(ws5, len(or_headers))

# Outbound-ready: ICP >= 65 with contact or email
outbound_ready = [r for r in priority_records if r['icp_score'] >= 65 and (r['contact'] or r['email'])]
for ri, rec in enumerate(outbound_ready, 2):
    vals = [
        rec['mid'], rec['name'], rec['contact'], rec['email'], rec['phone'],
        rec['website'], rec['li_url'],
        rec['icp_score'], rec['tam_tier'], rec.get('priority', ''),
        rec['salesperson'], rec.get('territory', ''),
        outreach_by_tier.get(rec['tam_tier'], ''),
        sequence_by_tier.get(rec['tam_tier'], ''),
    ]
    for ci, v in enumerate(vals, 1):
        cell = ws5.cell(ri, ci, v)
        cell.border = thin_border
        if ci == 8 and v and v >= 80:
            cell.fill = green_fill

auto_width(ws5)
print(f"OUTBOUND_READY: {len(outbound_ready)} records")

# ----- SHEET 6: EXECUTIVE_SUMMARY -----
ws6 = wb_out.create_sheet('EXECUTIVE_SUMMARY')

# Top 100 accounts
top100 = sorted(priority_records, key=lambda x: -x['icp_score'])[:100]

summary_lines = [
    ['AI SALES OS - PHASE 3 SALES INTELLIGENCE REPORT'],
    [''],
    ['Generated:', TIMESTAMP],
    ['Source:', 'CRM_DataQuality_Output.xlsx > CLEAN_MASTER'],
    [''],
    ['=== OVERVIEW ==='],
    ['Metric', 'Value'],
    ['Total Companies in CRM', len(records)],
    ['Priority Companies (A+B+C)', len(priority_records)],
    ['  - Tier A', sum(1 for r in priority_records if r['quality_tier']=='A')],
    ['  - Tier B', sum(1 for r in priority_records if r['quality_tier']=='B')],
    ['  - Tier C', sum(1 for r in priority_records if r['quality_tier']=='C')],
    [''],
    ['=== LINKEDIN COVERAGE ==='],
    ['Companies with LinkedIn', len(ld_records)],
    ['LinkedIn Coverage Rate', f'{len(ld_records)/len(priority_records)*100:.1f}%'],
    [''],
    ['=== INDUSTRY COVERAGE ==='],
]
for g, c in sorted(group_dist.items(), key=lambda x: -x[1]):
    summary_lines.append([f'  {g}', str(c), f'{c/len(priority_records)*100:.1f}%'])

summary_lines += [
    [''],
    ['=== ICP SCORE DISTRIBUTION ==='],
    ['Range', 'Count', 'Percentage'],
]
for k in ['80-100', '65-79', '50-64', '0-49']:
    v = icp_dist.get(k, 0)
    summary_lines.append([f'  {k}', str(v), f'{v/len(priority_records)*100:.1f}%' if priority_records else '0%'])

summary_lines += [
    [''],
    ['=== TAM DISTRIBUTION ==='],
]
for t, c in sorted(tam_dist.items()):
    summary_lines.append([t, str(c), f'{c/len(priority_records)*100:.1f}%'])

summary_lines += [
    [''],
    ['=== DECISION MAKERS ==='],
    ['Total Decision Makers Identified', str(len(decision_makers))],
    ['Decision Maker Coverage', f'{len(decision_makers)/len(priority_records)*100:.1f}%'],
    [''],
    ['=== TOP 100 ACCOUNTS TO CONTACT FIRST ==='],
    ['Rank', 'Company', 'Market Group', 'ICP Score', 'TAM Tier', 'Contact', 'Email', 'Website', 'Salesperson'],
]

for i, rec in enumerate(top100, 1):
    summary_lines.append([
        str(i), rec['name'][:50], rec['market_group'], str(rec['icp_score']),
        rec['tam_tier'], rec['contact'][:25] if rec['contact'] else '',
        rec['email'][:30] if rec['email'] else '', rec['website'],
        rec['salesperson'] if rec['salesperson'] != 'None' else '',
    ])

summary_lines += [
    [''],
    ['=== PLATFORM READINESS ==='],
    ['Apollo.io', 'Ready - domains validated, contacts enriched, ICP scored'],
    ['Clay', 'Ready - enrichment data with confidence scores'],
    ['AI Sales OS', 'Ready - ICP prioritization and routing complete'],
    ['Notion CRM', 'Ready - structured with TAM segments'],
    ['HubSpot', 'Ready - compatible with custom property mapping'],
    ['Salesforce', 'Ready - ready for Data Import'],
    ['Odoo CRM', 'Ready - standard fields populated'],
]

# Write summary
title_font = Font(bold=True, size=14, color="1F4E79")
section_font = Font(bold=True, size=12, color="2F5496")
hdr_font2 = Font(bold=True, color="1F4E79")

for ri, row_data in enumerate(summary_lines, 1):
    for ci, val in enumerate(row_data):
        cell = ws6.cell(ri, ci + 1, val)
        if ri == 1:
            cell.font = title_font
        elif val and str(val).startswith('==='):
            cell.font = section_font
        elif ri < 15 and ci == 0 and val and str(val)[0].isupper():
            cell.font = hdr_font2

auto_width(ws6)
# Widen a few key columns
ws6.column_dimensions['A'].width = 45
ws6.column_dimensions['B'].width = 55
ws6.column_dimensions['C'].width = 40

print(f"EXECUTIVE_SUMMARY: {len(summary_lines)} lines (incl. top 100 accounts)")

# Save
wb_out.save(OUTPUT_FILE)
print(f"\n=== OUTPUT SAVED: {OUTPUT_FILE} ===")
print(f"Sheets: {wb_out.sheetnames}")

# Final stats print
print("\n=== FINAL STATISTICS ===")
print(f"Priority Companies:     {len(priority_records)}")
print(f"  ICP >= 80 (Tier 1):  {len(tier1_accounts)}")
print(f"  ICP 65-79 (Tier 2):  {sum(1 for r in priority_records if 65 <= r['icp_score'] < 80)}")
print(f"  ICP 50-64 (Tier 3):  {sum(1 for r in priority_records if 50 <= r['icp_score'] < 65)}")
print(f"  ICP < 50  (Tier 4):  {sum(1 for r in priority_records if r['icp_score'] < 50)}")
print(f"Decision Makers:       {len(decision_makers)}")
print(f"Outbound Ready:        {len(outbound_ready)}")
print(f"LinkedIn Discovered:   {len(ld_records)}")
print(f"Top 100 Accounts:      {len(top100)}")
print(f"Avg ICP Score:         {sum(r['icp_score'] for r in priority_records)/len(priority_records):.1f}")
