"""Export scraped data to CSV, Excel, and JSON with full dynamic fields"""

import os
import csv
import json
from config import OUTPUT


def _collect_fieldnames(offices):
    fieldnames = []
    seen = set()
    for o in offices:
        for k in o:
            if k not in seen:
                seen.add(k)
                fieldnames.append(k)
    # Move common fields to front
    priority = ["office_id", "office_name", "hashed_office_id", "classification_grade", "latitude", "longitude"]
    ordered = [f for f in priority if f in seen]
    rest = sorted(f for f in fieldnames if f not in ordered)
    return ordered + rest


def _flatten(val):
    if isinstance(val, (list, dict)):
        return json.dumps(val, ensure_ascii=False)
    return val


def export_csv(offices, filepath=None):
    if filepath is None:
        filepath = OUTPUT["csv"]
    if not offices:
        return filepath
    fieldnames = _collect_fieldnames(offices)
    with open(filepath, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for office in offices:
            row = {k: _flatten(office.get(k, "")) for k in fieldnames}
            writer.writerow(row)
    return filepath


def export_xlsx(offices, filepath=None):
    if filepath is None:
        filepath = OUTPUT["xlsx"]
    if not offices:
        return filepath
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Engineering Offices"
        fieldnames = _collect_fieldnames(offices)
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="07706D", end_color="07706D", fill_type="solid")
        for col, fname in enumerate(fieldnames, 1):
            cell = ws.cell(row=1, column=col, value=fname)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        for row_idx, office in enumerate(offices, 2):
            for col_idx, fname in enumerate(fieldnames, 1):
                val = _flatten(office.get(fname, ""))
                ws.cell(row=row_idx, column=col_idx, value=val)
        ws.auto_filter.ref = f"A1:{chr(64 + len(fieldnames))}{len(offices) + 1}"
        wb.save(filepath)
    except ImportError:
        raise ImportError("openpyxl is required. Install: pip install openpyxl")
    return filepath


def export_json(offices, filepath=None):
    if filepath is None:
        filepath = OUTPUT["json"]
    with open(filepath, "w", encoding="utf-8") as f:
        json.dump(offices, f, ensure_ascii=False, indent=2)
    return filepath


def export_all(offices, full=False):
    results = {}
    if full:
        results["csv"] = export_csv(offices, "engineering_offices_full.csv")
        try:
            results["xlsx"] = export_xlsx(offices, "engineering_offices_full.xlsx")
        except ImportError:
            results["xlsx"] = None
        results["json"] = export_json(offices, "engineering_offices_full.json")
    else:
        results["csv"] = export_csv(offices)
        try:
            results["xlsx"] = export_xlsx(offices)
        except ImportError:
            results["xlsx"] = None
        results["json"] = export_json(offices)
    return results
