import openpyxl, sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
wb = openpyxl.load_workbook('Website_LinkedIn_Enrichment.xlsx')

# Check WEBSITE_NOT_FOUND
ws = wb['WEBSITE_NOT_FOUND']
print(f"WEBSITE_NOT_FOUND: {ws.max_row-1} records")
for r in range(2, min(35, ws.max_row+1)):
    mid = ws.cell(r,1).value
    name = str(ws.cell(r,2).value or '')[:40]
    domain = ws.cell(r,5).value or ''
    reason = str(ws.cell(r,4).value or '')[:60]
    priority = str(ws.cell(r,3).value or '')[:10]
    print(f"{mid} ({priority}): {name} | {domain} | {reason}")

# Check Tier 1 specific - compare with ICP sheet
print("\n=== TIER 1 WEBSITE STATUS ===")
wb2 = openpyxl.load_workbook('Website_LinkedIn_Enrichment.xlsx')
wb3 = openpyxl.load_workbook('Sales_Intelligence_Output_COPY.xlsx')
tier1 = wb3['ICP_PRIORITY_ACCOUNTS']

for t in range(2, tier1.max_row+1):
    mid = tier1.cell(t,1).value
    name = str(tier1.cell(t,2).value or '')[:50]
    
    # Find in results
    for ws_name in ['WEBSITE_VERIFIED', 'WEBSITE_NOT_FOUND']:
        ws_check = wb2[ws_name]
        for r in range(2, ws_check.max_row+1):
            if ws_check.cell(r,1).value == mid:
                status = 'VERIFIED' if ws_name == 'WEBSITE_VERIFIED' else ws_check.cell(r,4).value
                domain = ws_check.cell(r,5).value if ws_name == 'WEBSITE_NOT_FOUND' else ''
                notes = str(ws_check.cell(r,7 if ws_name == 'WEBSITE_VERIFIED' else 4).value or '')[:60]
                li_found = ''
                wb2_li = wb2['LINKEDIN_VERIFIED']
                for lr in range(2, wb2_li.max_row+1):
                    if wb2_li.cell(lr,1).value == mid:
                        li_found = 'YES'
                if not li_found:
                    li_found = 'NO'
                print(f"{mid}: {name} | WS={status} | Domain={domain} | LI={li_found}")
                break
