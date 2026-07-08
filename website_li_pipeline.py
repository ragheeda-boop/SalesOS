"""
Website & LinkedIn First Enrichment Pipeline
Focus: Discover + Validate only. No ICP, no outreach.
"""
import openpyxl, json, re, sys, io, socket, time, ssl
from datetime import datetime
from urllib.request import Request, urlopen
from urllib.error import URLError, HTTPError
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from collections import Counter

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

TODAY = datetime.now().strftime('%Y-%m-%d')
TIMESTAMP = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

hdr_font = Font(bold=True, color="FFFFFF", size=10)
hdr_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
hdr_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                      top=Side(style='thin'), bottom=Side(style='thin'))
green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

def style_header(ws, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(1, c)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = hdr_align
        cell.border = thin_border

def auto_width(ws):
    for col_cells in ws.columns:
        max_len = 0
        col_letter = openpyxl.utils.get_column_letter(col_cells[0].column)
        for cell in col_cells:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except: pass
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 12), 70)

def extract_domain(url):
    if not url: return ''
    m = re.search(r'https?://(?:www\.)?([^/]+)', str(url))
    return m.group(1).lower() if m else ''

def check_domain_http(domain):
    """Check if domain resolves and HTTP loads, return status + title."""
    result = {'resolves': False, 'http_ok': False, 'title': '', 'error': '', 'redirect': '', 'description': ''}
    if not domain: return result
    
    try:
        socket.getaddrinfo(domain, 80, socket.AF_INET, socket.SOCK_STREAM)
        result['resolves'] = True
    except:
        result['error'] = 'DNS failed'
        return result
    
    for proto in ['https', 'http']:
        try:
            ctx = ssl.create_default_context()
            ctx.check_hostname = False
            ctx.verify_mode = ssl.CERT_NONE
            req = Request(f'{proto}://{domain}', headers={
                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36',
                'Accept': 'text/html',
                'Accept-Language': 'en-US,en;q=0.9',
            })
            resp = urlopen(req, timeout=10, context=ctx)
            result['http_ok'] = resp.status == 200
            result['redirect'] = resp.url if resp.url.lower().rstrip('/') != f'{proto}://{domain}'.lower().rstrip('/') and resp.url.lower().rstrip('/') != f'{proto}://www.{domain}'.lower().rstrip('/') else ''
            
            content = resp.read(131072).decode('utf-8', errors='replace')
            # Extract title
            m = re.search(r'<title[^>]*>(.*?)</title>', content, re.IGNORECASE | re.DOTALL)
            if m: result['title'] = m.group(1).strip()[:200]
            
            # Extract meta description
            m = re.search(r'<meta[^>]+name=["\']description["\'][^>]+content=["\']([^"\']+)["\']', content, re.IGNORECASE)
            if m: result['description'] = m.group(1).strip()[:300]
            
            # Check for parked indicators
            parked_kw = ['godaddy', 'sedo', 'hugedomains', 'domain parking', 'buy this domain', 'this domain is parked', 'domain is for sale']
            title_lower = (result['title'] or '').lower()
            if any(kw in title_lower for kw in parked_kw):
                result['parked'] = True
                result['error'] = f'Parked domain: {result["title"][:100]}'
            else:
                result['parked'] = False
            break
        except HTTPError as e:
            if e.code == 404:
                continue
            result['error'] = f'HTTP {e.code}'
        except URLError as e:
            result['error'] = f'URL Error: {e.reason}'
        except Exception as e:
            result['error'] = str(e)[:100]
    
    return result

# ===================== LOAD DATA =====================
print("=== LOADING INPUT ===")
src_path = 'Sales_Intelligence_Output_COPY.xlsx'
wb_src = openpyxl.load_workbook(src_path)

# Load Tier 1 accounts
ws1 = wb_src['ICP_PRIORITY_ACCOUNTS']
S1 = {}
for c in range(1, ws1.max_column + 1):
    S1[ws1.cell(1, c).value] = c

print(f"ICP_PRIORITY_ACCOUNTS: {ws1.max_row - 1} records")

# Build unified company list from ICP_PRIORITY first, then SALES_INTELLIGENCE_MASTER
companies = []
seen_mids = set()

for r in range(2, ws1.max_row + 1):
    mid = str(ws1.cell(r, 1).value or '')
    if mid in seen_mids: continue
    seen_mids.add(mid)
    companies.append({
        'mid': mid,
        'name': str(ws1.cell(r, 2).value or ''),
        'priority': 1,
        'existing_website': str(ws1.cell(r, S1.get('Website', 7)).value or ''),
        'existing_li': str(ws1.cell(r, S1.get('LinkedIn URL', 8)).value or ''),
        'existing_domain': extract_domain(str(ws1.cell(r, S1.get('Website', 7)).value or '')),
    })

# Add remaining from SALES_INTELLIGENCE_MASTER (non-Tier 1)
ws2 = wb_src['SALES_INTELLIGENCE_MASTER']
S2 = {}
for c in range(1, ws2.max_column + 1):
    S2[ws2.cell(1, c).value] = c

print(f"SALES_INTELLIGENCE_MASTER: {ws2.max_row - 1} records")

for r in range(2, ws2.max_row + 1):
    mid = str(ws2.cell(r, 1).value or '')
    if mid in seen_mids: continue
    seen_mids.add(mid)
    companies.append({
        'mid': mid,
        'name': str(ws2.cell(r, 2).value or ''),
        'priority': 2,
        'existing_website': str(ws2.cell(r, S2.get('Validated Website', 9)).value or ''),
        'existing_li': str(ws2.cell(r, S2.get('LinkedIn URL', 12)).value or ''),
        'existing_domain': extract_domain(str(ws2.cell(r, S2.get('Validated Website', 9)).value or '')),
    })

print(f"Total unique companies to process: {len(companies)}")
print(f"  Tier 1: {sum(1 for c in companies if c['priority']==1)}")
print(f"  Tier 2-4: {sum(1 for c in companies if c['priority']==2)}")

# Phase 1: Validate ALL existing websites via HTTP check
print(f"\n=== DOMAIN VALIDATION ===")
# Also load known LinkedIn discoveries from earlier
known_li = {
    'COMP-000133': 'https://sa.linkedin.com/company/megaminditsolutions',
    'COMP-000718': 'https://www.linkedin.com/company/lasetech-ksa',
    'COMP-000069': 'https://www.linkedin.com/company/altaawin-medical-group',
    'COMP-000084': 'https://sa.linkedin.com/company/ratl-technology',
    'COMP-000043': 'https://sa.linkedin.com/company/mohrkey',
    'COMP-000062': 'https://sa.linkedin.com/company/zincoint',
    'COMP-000050': 'https://sa.linkedin.com/company/seven-eleven-trading',
    'COMP-000033': 'https://sa.linkedin.com/company/creativeteamforce',
    'COMP-000029': 'https://www.linkedin.com/company/sedres-maritime-co-ltd',
    'COMP-000025': 'https://www.linkedin.com/company/nama-arabia-trading-company',
    'COMP-000009': 'https://www.linkedin.com/company/bahasan-trading',
}

known_li_conf = {
    'COMP-000133': 100, 'COMP-000718': 100, 'COMP-000069': 85, 'COMP-000084': 70,
    'COMP-000043': 100, 'COMP-000062': 100, 'COMP-000050': 90,
    'COMP-000033': 85, 'COMP-000029': 100, 'COMP-000025': 85, 'COMP-000009': 70,
}

results = []
count = 0
for c in companies:
    count += 1
    mid = c['mid']
    domain = c['existing_domain']
    website = c['existing_website']
    
    rec = {
        'mid': mid,
        'name': c['name'],
        'priority': 'Tier 1' if c['priority'] == 1 else 'Tier 2-4',
        'website': website,
        'domain': domain,
        'website_conf': 0,
        'website_status': 'NOT_FOUND',
        'website_notes': '',
        'li_url': '',
        'li_conf': 0,
        'li_status': 'NOT_FOUND',
        'li_notes': '',
        'company_desc': '',
        'phone': '',
        'email': '',
        'city': '',
        'country': 'Saudi Arabia',
    }
    
    # Apply known LinkedIn
    if mid in known_li:
        rec['li_url'] = known_li[mid]
        rec['li_conf'] = known_li_conf.get(mid, 85)
        rec['li_status'] = 'VERIFIED' if rec['li_conf'] >= 100 else 'CONFIRMED'
    
    # Validate existing domain
    if domain:
        check = check_domain_http(domain)
        
        if check.get('parked'):
            rec['website_status'] = 'PARKED'
            rec['website_conf'] = 0
            rec['website_notes'] = check.get('error', 'Domain appears parked')
        elif check['http_ok']:
            rec['website_status'] = 'VERIFIED'
            rec['website_conf'] = 85  # Base - needs name match for 100
            rec['website_notes'] = f"HTTP 200 | Title: {(check['title'] or '')[:150]}"
            rec['company_desc'] = check.get('description', '')
            
            # Check if company name appears in title
            name_parts = c['name'].replace('شركة', '').replace('مؤسسة', '').replace('Company', '').replace('Ltd', '').strip()[:30]
            title_lower = (check['title'] or '').lower()
            if name_parts.lower()[:15] in title_lower or any(p.lower() in title_lower for p in name_parts.split() if len(p) > 3):
                rec['website_conf'] = 100
                rec['website_notes'] += ' | Name match confirmed'
        elif check['resolves']:
            rec['website_status'] = 'DNS_ONLY'
            rec['website_conf'] = 25
            rec['website_notes'] = f"DNS resolves but HTTP error: {check.get('error', 'Unknown')}"
        else:
            rec['website_status'] = 'FAILED'
            rec['website_conf'] = 0
            rec['website_notes'] = f"DNS failed: {check.get('error', 'Unknown')}"
    else:
        rec['website_status'] = 'NOT_FOUND'
        rec['website_notes'] = 'No website in source data'
    
    results.append(rec)
    if count % 30 == 0:
        print(f"  Checked {count}/{len(companies)}...")

# Summary
ws_verified = [r for r in results if r['website_status'] == 'VERIFIED']
ws_parked = [r for r in results if r['website_status'] == 'PARKED']
ws_dns = [r for r in results if r['website_status'] == 'DNS_ONLY']
ws_failed = [r for r in results if r['website_status'] == 'FAILED']
ws_notfound = [r for r in results if r['website_status'] == 'NOT_FOUND']

li_verified = [r for r in results if r['li_status'] in ('VERIFIED', 'CONFIRMED')]

print(f"\n=== VALIDATION RESULTS ===")
print(f"Website VERIFIED:  {len(ws_verified)} ({len(ws_verified)/len(results)*100:.1f}%)")
print(f"Website PARKED:    {len(ws_parked)}")
print(f"Website DNS_ONLY:  {len(ws_dns)}")
print(f"Website FAILED:    {len(ws_failed)}")
print(f"Website NOT_FOUND: {len(ws_notfound)}")
print(f"Total websites:    {len(ws_verified) + len(ws_parked)} (verified + parked)")
print(f"")
print(f"LinkedIn VERIFIED: {len(li_verified)} ({len(li_verified)/len(results)*100:.1f}%)")

# ===================== GENERATE OUTPUT =====================
print(f"\n=== GENERATING OUTPUT ===")
OUTPUT_FILE = 'Website_LinkedIn_Enrichment.xlsx'
wb_out = openpyxl.Workbook()

# --- SHEET 1: WEBSITE_VERIFIED ---
ws_out1 = wb_out.active
ws_out1.title = 'WEBSITE_VERIFIED'
wv_headers = ['Company Master ID', 'Company Name', 'Priority', 'Website', 'Domain',
              'Confidence', 'Status', 'Title', 'Description', 'Phone', 'Email',
              'City', 'Country', 'Last Verified']
for i, h in enumerate(wv_headers, 1):
    ws_out1.cell(1, i, h)
style_header(ws_out1, len(wv_headers))

for ri, r in enumerate(ws_verified, 2):
    vals = [r['mid'], r['name'], r['priority'], r['website'], r['domain'],
            r['website_conf'], r['website_status'], r['company_desc'][:100], r['company_desc'],
            r['phone'], r['email'], r['city'], r['country'], TODAY]
    for ci, v in enumerate(vals, 1):
        c = ws_out1.cell(ri, ci, v)
        c.border = thin_border
        if ci == 6 and v and v >= 100:
            c.fill = green_fill

auto_width(ws_out1)
print(f"WEBSITE_VERIFIED: {len(ws_verified)} records")

# --- SHEET 2: LINKEDIN_VERIFIED ---
ws_out2 = wb_out.create_sheet('LINKEDIN_VERIFIED')
lv_headers = ['Company Master ID', 'Company Name', 'Priority', 'LinkedIn URL',
              'Confidence', 'Status', 'Employees', 'Followers',
              'Industry', 'Description', 'Notes', 'Last Verified']
for i, h in enumerate(lv_headers, 1):
    ws_out2.cell(1, i, h)
style_header(ws_out2, len(lv_headers))

for ri, r in enumerate(li_verified, 2):
    vals = [r['mid'], r['name'], r['priority'], r['li_url'],
            r['li_conf'], r['li_status'], '', '',
            '', '', r['li_notes'], TODAY]
    for ci, v in enumerate(vals, 1):
        c = ws_out2.cell(ri, ci, v)
        c.border = thin_border
        if ci == 5 and v and v >= 100:
            c.fill = green_fill

auto_width(ws_out2)
print(f"LINKEDIN_VERIFIED: {len(li_verified)} records")

# --- SHEET 3: WEBSITE_NOT_FOUND ---
ws_out3 = wb_out.create_sheet('WEBSITE_NOT_FOUND')
wnf_headers = ['Company Master ID', 'Company Name', 'Priority', 'Reason', 'Domain Tried']
for i, h in enumerate(wnf_headers, 1):
    ws_out3.cell(1, i, h)
style_header(ws_out3, len(wnf_headers))

website_not_found = ws_parked + ws_dns + ws_failed + ws_notfound
for ri, r in enumerate(website_not_found, 2):
    vals = [r['mid'], r['name'], r['priority'], r['website_notes'], r['domain']]
    for ci, v in enumerate(vals, 1):
        c = ws_out3.cell(ri, ci, v)
        c.border = thin_border

auto_width(ws_out3)
print(f"WEBSITE_NOT_FOUND: {len(website_not_found)} records")

# --- SHEET 4: LINKEDIN_NOT_FOUND ---
ws_out4 = wb_out.create_sheet('LINKEDIN_NOT_FOUND')
lnf_headers = ['Company Master ID', 'Company Name', 'Priority', 'Website Status', 'Notes']
for i, h in enumerate(lnf_headers, 1):
    ws_out4.cell(1, i, h)
style_header(ws_out4, len(lnf_headers))

li_not_found = [r for r in results if r['li_status'] == 'NOT_FOUND']
for ri, r in enumerate(li_not_found, 2):
    vals = [r['mid'], r['name'], r['priority'], r['website_status'], 'LinkedIn not discovered yet']
    for ci, v in enumerate(vals, 1):
        c = ws_out4.cell(ri, ci, v)
        c.border = thin_border

auto_width(ws_out4)
print(f"LINKEDIN_NOT_FOUND: {len(li_not_found)} records")

# --- SHEET 5: HIGH_CONFIDENCE_MATCHES ---
ws_out5 = wb_out.create_sheet('HIGH_CONFIDENCE_MATCHES')
hc_headers = ['Company Master ID', 'Company Name', 'Website', 'Confidence',
              'LinkedIn URL', 'LI Confidence', 'Priority', 'Notes']
for i, h in enumerate(hc_headers, 1):
    ws_out5.cell(1, i, h)
style_header(ws_out5, len(hc_headers))

high_conf = [r for r in results if r['website_conf'] >= 100 or r['li_conf'] >= 100]
for ri, r in enumerate(high_conf, 2):
    vals = [r['mid'], r['name'], r['website'], r['website_conf'],
            r['li_url'], r['li_conf'], r['priority'],
            f"WS: {r['website_status']} | LI: {r['li_status']}"]
    for ci, v in enumerate(vals, 1):
        c = ws_out5.cell(ri, ci, v)
        c.border = thin_border
        if ci in (4, 6) and v and v >= 100:
            c.fill = green_fill

auto_width(ws_out5)
print(f"HIGH_CONFIDENCE_MATCHES: {len(high_conf)} records")

# --- SHEET 6: EXECUTIVE_SUMMARY ---
ws_out6 = wb_out.create_sheet('EXECUTIVE_SUMMARY')

tot = len(results)
ws_ok = len(ws_verified)
li_ok = len(li_verified)
tier1 = [r for r in results if r['priority'] == 'Tier 1']
tier1_ws = len([r for r in tier1 if r['website_status'] == 'VERIFIED'])
tier1_li = len([r for r in tier1 if r['li_status'] in ('VERIFIED', 'CONFIRMED')])

summary = [
    ['WEBSITE & LINKEDIN ENRICHMENT REPORT'],
    [''],
    ['Generated:', TIMESTAMP],
    ['Source:', 'Sales_Intelligence_Output.xlsx'],
    [''],
    ['=== COVERAGE OVERVIEW ==='],
    ['Metric', 'Total', 'Tier 1 Only'],
    ['Companies', str(tot), str(len(tier1))],
    ['Website Verified', f"{ws_ok} ({ws_ok/tot*100:.1f}%)", f"{tier1_ws} ({tier1_ws/len(tier1)*100:.1f}%)" if tier1 else '0'],
    ['LinkedIn Verified', f"{li_ok} ({li_ok/tot*100:.1f}%)", f"{tier1_li} ({tier1_li/len(tier1)*100:.1f}%)" if tier1 else '0'],
    ['High Confidence (100%)', str(len(high_conf)), ''],
    [''],
    ['=== WEBSITE STATUS DISTRIBUTION ==='],
    ['Status', 'Count', 'Percentage'],
    ['VERIFIED', str(len(ws_verified)), f"{len(ws_verified)/tot*100:.1f}%"],
    ['PARKED', str(len(ws_parked)), f"{len(ws_parked)/tot*100:.1f}%"],
    ['DNS_ONLY', str(len(ws_dns)), f"{len(ws_dns)/tot*100:.1f}%"],
    ['FAILED', str(len(ws_failed)), f"{len(ws_failed)/tot*100:.1f}%"],
    ['NOT_FOUND', str(len(ws_notfound)), f"{len(ws_notfound)/tot*100:.1f}%"],
    [''],
    ['=== SUCCESS CRITERIA CHECK ==='],
    ['Criteria', 'Target', 'Current', 'Status'],
    ['Website Coverage', '> 70%', f'{ws_ok/tot*100:.1f}%', 'PASS' if ws_ok/tot >= 0.70 else 'NEEDS WORK'],
    ['LinkedIn Coverage', '> 50%', f'{li_ok/tot*100:.1f}%', 'PASS' if li_ok/tot >= 0.50 else 'NEEDS WORK'],
    ['Confidence Score >= 85', 'All verified', f'{len(high_conf)} at 100%', 'IN PROGRESS'],
    [''],
    ['=== RECOMMENDATIONS ==='],
    ['1. Run targeted LinkedIn searches for the 29 Tier 1 accounts (only {tier1_li} found so far)'],
    ['2. Verify parked domains ({len(ws_parked)}) - may have moved to new URLs'],
    ['3. Retry DNS_ONLY domains ({len(ws_dns)}) with browser-based check'],
    ['4. Manual name-match verification for the {len(ws_verified)} VERIFIED websites'],
]

title_font = Font(bold=True, size=14, color="1F4E79")
section_font = Font(bold=True, size=12, color="2F5496")

for ri, row in enumerate(summary, 1):
    for ci, val in enumerate(row):
        cell = ws_out6.cell(ri, ci + 1, val)
        if ri == 1: cell.font = title_font
        if val and str(val).startswith('==='): cell.font = section_font

auto_width(ws_out6)
ws_out6.column_dimensions['A'].width = 40
ws_out6.column_dimensions['B'].width = 30
ws_out6.column_dimensions['C'].width = 30

# Save
wb_out.save(OUTPUT_FILE)
print(f"\n=== OUTPUT SAVED: {OUTPUT_FILE} ===")
print(f"Sheets: {wb_out.sheetnames}")
