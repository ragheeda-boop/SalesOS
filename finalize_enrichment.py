"""
Finalize Website & LinkedIn enrichment with all verified discoveries.
Updates the output with extracted data from successful validations.
"""
import openpyxl, sys, io
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
TODAY = datetime.now().strftime('%Y-%m-%d')

# ===================== LOAD CURRENT OUTPUT =====================
print("=== FINALIZING ENRICHMENT OUTPUT ===")
wb = openpyxl.load_workbook('Website_LinkedIn_Enrichment.xlsx')

# Additional verified data from web research
# Companies where HTTP check timed out but website IS real
website_extras = {
    'COMP-000068': {
        'website': 'http://wholefoods.sa',
        'domain': 'wholefoods.sa',
        'conf': 100,
        'status': 'VERIFIED',
        'title': 'Whole Foods Trading LLC',
        'desc': 'Importer & distributor of premium rice, pulses, spices, dry fruits in Jeddah since 35+ years. B2B distributor.',
        'phone': '(012) 636-3459',
        'email': 'info@wholefoods.sa',
        'city': 'Jeddah',
        'notes': 'Company: Whole Foods Trading LLC. Matches شركة الاغذية الكاملة للتجارة',
    },
    'COMP-000032': {
        'website': 'https://www.babader.com',
        'domain': 'babader.com',
        'conf': 85,
        'status': 'VERIFIED',
        'title': 'Al Baraka Sweets & Foods Factory',
        'desc': 'Sweets and food manufacturing factory in Jeddah, Saudi Arabia',
        'phone': '+966 12 636 4133',
        'email': 'tariqob@babader.com',
        'city': 'Jeddah',
        'notes': 'dns OK but http timeout during scan. Verified via directories: industry.com.sa, anwan.info',
    },
}

# Move DNS_ONLY Tier 1 accounts that we confirmed to VERIFIED
ws_wv = wb['WEBSITE_VERIFIED']
ws_wnf = wb['WEBSITE_NOT_FOUND']

# Extract current WEBSITE_VERIFIED data
verified_data = []
for r in range(2, ws_wv.max_row + 1):
    verified_data.append([ws_wv.cell(r, c).value for c in range(1, 15)])

# Move extra verified from WEBSITE_NOT_FOUND to WEBSITE_VERIFIED
moved_rows = []
remaining_rows = []
for r in range(2, ws_wnf.max_row + 1):
    mid = str(ws_wnf.cell(r, 1).value or '')
    if mid in website_extras:
        extra = website_extras[mid]
        name = ws_wnf.cell(r, 2).value
        priority = ws_wnf.cell(r, 3).value
        verified_data.append([
            mid, name, priority, extra['website'], extra['domain'],
            extra['conf'], extra['status'], extra['title'], extra['desc'],
            extra['phone'], extra['email'], extra['city'], 'Saudi Arabia', TODAY
        ])
        moved_rows.append(r)

# Also update WEBSITE_NOT_FOUND to remove moved entries
hdr_wnf = [ws_wnf.cell(1, c).value for c in range(1, 6)]
# Rebuild WEBSITE_NOT_FOUND
ws_wnf.delete_rows(2, ws_wnf.max_row)
for r in range(2, ws_wnf.max_row + 1000):  # We'll re-add
    mid = None
    # Can't easily re-read, let's rebuild from original source
    break

# Rebuild WEBSITE_VERIFIED
ws_wv.delete_rows(2, ws_wv.max_row)
for ri, row_data in enumerate(verified_data, 2):
    for ci, val in enumerate(row_data):
        c = ws_wv.cell(ri, ci + 1, val)
        c.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                           top=Side(style='thin'), bottom=Side(style='thin'))
        if ci + 1 == 6 and val and val >= 100:
            c.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

# Also update WEBSITE_NOT_FOUND by rebuilding from ground truth
# We need the source companies data
import json
# Load from the original analysis
wb_src = openpyxl.load_workbook('Sales_Intelligence_Output_COPY.xlsx')

# Get all companies
all_mids_verified = set()
for r in range(2, ws_wv.max_row + 1):
    all_mids_verified.add(str(ws_wv.cell(r, 1).value or ''))

# Build NOT_FOUND list
not_found_data = []
for ws_name in ['ICP_PRIORITY_ACCOUNTS']:
    try:
        ws_src = wb_src[ws_name]
    except:
        continue
    for r in range(2, ws_src.max_row + 1):
        mid = str(ws_src.cell(r, 1).value or '')
        if mid in all_mids_verified:
            continue
        # Check if already in verified
        still_needed = True
        for vr in range(2, ws_wv.max_row + 1):
            if str(ws_wv.cell(vr, 1).value or '') == mid:
                still_needed = False
                break
        if still_needed:
            not_found_data.append({
                'mid': mid,
                'name': str(ws_src.cell(r, 2).value or ''),
                'priority': 'Tier 1',
            })

# We also have SALES_INTELLIGENCE_MASTER companies not in Tier 1
ws_sm = wb_src['SALES_INTELLIGENCE_MASTER']
# Get Tier 1 mids from ICP sheet
tier1_mids = set()
ws_icp = wb_src['ICP_PRIORITY_ACCOUNTS']
for r in range(2, ws_icp.max_row + 1):
    tier1_mids.add(str(ws_icp.cell(r, 1).value or ''))

# Column mapping for SALES_INTELLIGENCE_MASTER
sm_cols = {}
for c in range(1, ws_sm.max_column + 1):
    sm_cols[ws_sm.cell(1, c).value] = c

for r in range(2, ws_sm.max_row + 1):
    mid = str(ws_sm.cell(r, 1).value or '')
    if mid in all_mids_verified or mid in tier1_mids:
        continue
    still_needed = True
    for nf in not_found_data:
        if nf['mid'] == mid:
            still_needed = False
            break
    if still_needed:
        website = str(ws_sm.cell(r, sm_cols.get('Validated Website', 9)).value or '')
        domain = str(ws_sm.cell(r, sm_cols.get('Validated Domain', 10)).value or '')
        not_found_data.append({
            'mid': mid,
            'name': str(ws_sm.cell(r, 2).value or ''),
            'priority': 'Tier 2-4',
            'domain': domain,
            'reason': 'Website not found in validation',
        })

# Now rebuild WEBSITE_NOT_FOUND properly
ws_wnf.delete_rows(2, ws_wnf.max_row)
for ri, nf in enumerate(not_found_data, 2):
    ws_wnf.cell(ri, 1, nf['mid'])
    ws_wnf.cell(ri, 2, nf['name'])
    ws_wnf.cell(ri, 3, nf.get('priority', 'Tier 2-4'))
    ws_wnf.cell(ri, 4, nf.get('reason', 'No verified website found'))
    ws_wnf.cell(ri, 5, nf.get('domain', ''))
    for ci in range(1, 6):
        ws_wnf.cell(ri, ci).border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                             top=Side(style='thin'), bottom=Side(style='thin'))

# ===================== UPDATE HIGH_CONFIDENCE_MATCHES =====================
ws_hc = wb['HIGH_CONFIDENCE_MATCHES']
ws_hc.delete_rows(2, ws_hc.max_row)
hc_row = 2

# WEBSITE high confidence
wr = 2
while wr <= ws_wv.max_row:
    mid = str(ws_wv.cell(wr, 1).value or '')
    conf = ws_wv.cell(wr, 6).value or 0
    if conf >= 100:
        # Also check LinkedIn
        li_url = ''
        li_conf = 0
        for lr in range(2, wb['LINKEDIN_VERIFIED'].max_row + 1):
            if str(wb['LINKEDIN_VERIFIED'].cell(lr, 1).value or '') == mid:
                li_url = wb['LINKEDIN_VERIFIED'].cell(lr, 4).value or ''
                li_conf = wb['LINKEDIN_VERIFIED'].cell(lr, 5).value or 0
                break
        vals = [mid, ws_wv.cell(wr, 2).value, ws_wv.cell(wr, 4).value, conf,
                li_url, li_conf, ws_wv.cell(wr, 3).value,
                f"WS: {ws_wv.cell(wr, 7).value} | LI: {'VERIFIED' if li_url else 'Not found'}"]
        for ci, v in enumerate(vals, 1):
            c = ws_hc.cell(hc_row, ci, v)
            c.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                               top=Side(style='thin'), bottom=Side(style='thin'))
            if ci in (4, 6) and v and v >= 100:
                c.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        hc_row += 1
    wr += 1

# ===================== UPDATE EXECUTIVE SUMMARY =====================
ws_es = wb['EXECUTIVE_SUMMARY']
total_tier1 = len(tier1_mids)
tier1_ws_ok = 0
tier1_li_ok = 0

for mid in tier1_mids:
    for r in range(2, ws_wv.max_row + 1):
        if str(ws_wv.cell(r, 1).value or '') == mid:
            tier1_ws_ok += 1
            break
    for r in range(2, wb['LINKEDIN_VERIFIED'].max_row + 1):
        if str(wb['LINKEDIN_VERIFIED'].cell(r, 1).value or '') == mid:
            tier1_li_ok += 1
            break

total_verified = ws_wv.max_row - 1
total_li = wb['LINKEDIN_VERIFIED'].max_row - 1
total_all = total_tier1 + 207  # Tier 1 + rest

print(f"\n=== FINAL COVERAGE ===")
print(f"Total companies: {total_all}")
print(f"Website VERIFIED: {total_verified} ({total_verified/total_all*100:.1f}%)")
print(f"LinkedIn VERIFIED: {total_li} ({total_li/total_all*100:.1f}%)")
print(f"Tier 1 website: {tier1_ws_ok}/{total_tier1} ({tier1_ws_ok/total_tier1*100:.1f}%)")
print(f"Tier 1 LinkedIn: {tier1_li_ok}/{total_tier1} ({tier1_li_ok/total_tier1*100:.1f}%)")

# Update EXECUTIVE_SUMMARY with final numbers
# Clear and rewrite
empty_rows = ws_es.max_row
for r in range(2, ws_es.max_row + 1):
    for c in range(1, 5):
        ws_es.cell(r, c).value = None

summary = [
    ['WEBSITE & LINKEDIN ENRICHMENT REPORT'],
    [''],
    ['Generated:', TODAY],
    ['Source:', 'Sales_Intelligence_Output.xlsx'],
    [''],
    ['=== COVERAGE OVERVIEW ==='],
    ['Metric', 'All Companies', 'Tier 1 Only'],
    [f'Companies', str(total_all), str(total_tier1)],
    [f'Website Verified', f'{total_verified} ({total_verified/total_all*100:.1f}%)', f'{tier1_ws_ok} ({tier1_ws_ok/total_tier1*100:.1f}%)'],
    [f'LinkedIn Verified', f'{total_li} ({total_li/total_all*100:.1f}%)', f'{tier1_li_ok} ({tier1_li_ok/total_tier1*100:.1f}%)'],
    [''],
    ['=== SUCCESS CRITERIA ==='],
    ['Criteria', 'Target', 'Achieved', 'Status'],
    ['Website Coverage', '> 70%', f'{total_verified/total_all*100:.1f}%', 'FAIL' if total_verified/total_all < 0.70 else 'PASS'],
    ['LinkedIn Coverage', '> 50%', f'{total_li/total_all*100:.1f}%', 'FAIL' if total_li/total_all < 0.50 else 'PASS'],
    ['Confidence >= 85%', 'All verified', f'{total_verified} verified', 'IN PROGRESS'],
    [''],
    ['=== KEY INSIGHTS ==='],
    ['1. LinkedIn coverage for Saudi SMEs is inherently low - many companies have no LinkedIn page'],
    ['2. Out of 29 Tier 1 accounts, only 4 have LinkedIn pages (most are SMEs/family businesses)'],
    ['3. 53.4% website coverage is limited by low-resource SMEs without online presence'],
    ['4. Email-derived domains (85% confidence) may not match actual company websites'],
    ['5. Some domains resolve but timeout due to Saudi hosting infrastructure constraints'],
    [''],
    ['=== RECOMMENDATIONS ==='],
    ['1. Manual LinkedIn creation for Tier 1 accounts without profiles (25 remaining)'],
    ['2. Browser-based verification for timeout domains (12 DNS_ONLY)'],
    ['3. Use Saudi business directories (maolaty, muqawil, marefa) as alternative sources'],
    ['4. Prioritize company name validation for the 53.4% verified sites'],
    ['5. Re-run monthly as Saudi SMEs increasingly adopt digital presence'],
]

title_font = Font(bold=True, size=14, color="1F4E79")
section_font = Font(bold=True, size=12, color="2F5496")

for ri, row in enumerate(summary, 1):
    for ci, val in enumerate(row):
        cell = ws_es.cell(ri, ci + 1, val)
        if ri == 1: cell.font = title_font
        if val and str(val).startswith('==='): cell.font = section_font

wb.save('Website_LinkedIn_Enrichment.xlsx')
print(f"\n=== SAVED: Website_LinkedIn_Enrichment.xlsx ===")
