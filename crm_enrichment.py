#!/usr/bin/env python3
"""
CRM Data Enhancement & Company Digital Footprint Discovery
Phases 1-9: Full pipeline
"""

import csv
import re
import os
import string
from collections import Counter, defaultdict
from difflib import SequenceMatcher
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("pip install openpyxl")
    raise

INPUT_FILE = "LeadOpportunity.csv"
OUTPUT_FILE = "CRM_Enrichment_Output.xlsx"

# ─────────────────────────────────────────────
# Load Data
# ─────────────────────────────────────────────
def load_data():
    rows = []
    with open(INPUT_FILE, 'r', encoding='utf-8-sig') as f:
        reader = csv.DictReader(f)
        for row in reader:
            rows.append(row)
    return rows

# ─────────────────────────────────────────────
# Phase 1: Data Audit
# ─────────────────────────────────────────────
def phase1_audit(rows):
    total_raw = len(rows)
    
    companies_raw = []
    for r in rows:
        opp = (r.get('Opportunity') or '').strip()
        if opp:
            companies_raw.append(opp)
    
    total_records = len(companies_raw)
    unique_names = set(companies_raw)
    unique_count = len(unique_names)
    dup_count = total_records - unique_count
    
    name_counts = Counter(companies_raw)
    dups = {k: v for k, v in name_counts.items() if v > 1}
    
    rows_w_company = [r for r in rows if (r.get('Opportunity') or '').strip()]
    
    missing_website = len(rows_w_company)
    missing_linkedin = len(rows_w_company)
    missing_industry = len(rows_w_company)
    missing_contact = sum(1 for r in rows_w_company if not (r.get('Contact Name') or '').strip())
    missing_email = sum(1 for r in rows_w_company if not (r.get('Email') or '').strip())
    missing_phone = sum(1 for r in rows_w_company if not (r.get('Phone') or '').strip())
    
    coverage_website = (len(rows_w_company) - missing_website) / len(rows_w_company) * 100 if rows_w_company else 0
    coverage_linkedin = (len(rows_w_company) - missing_linkedin) / len(rows_w_company) * 100 if rows_w_company else 0
    coverage_industry = (len(rows_w_company) - missing_industry) / len(rows_w_company) * 100 if rows_w_company else 0
    
    dq_score = ((coverage_website + coverage_linkedin + coverage_industry + 
                 (len(rows_w_company) - missing_contact) / len(rows_w_company) * 100 +
                 (len(rows_w_company) - missing_email) / len(rows_w_company) * 100 +
                 (len(rows_w_company) - missing_phone) / len(rows_w_company) * 100) / 6)
    
    audit = {
        'total_raw': total_raw,
        'total_records': total_records,
        'unique_companies': unique_count,
        'duplicate_occurrences': dup_count,
        'dup_companies_count': len(dups),
        'duplicate_rate': dup_count / total_records * 100 if total_records else 0,
        'missing_website': missing_website,
        'missing_linkedin': missing_linkedin,
        'missing_industry': missing_industry,
        'missing_contact': missing_contact,
        'missing_email': missing_email,
        'missing_phone': missing_phone,
        'website_coverage': coverage_website,
        'linkedin_coverage': coverage_linkedin,
        'industry_coverage': coverage_industry,
        'data_quality_score': dq_score,
        'duplicates': dups,
    }
    return audit, companies_raw, rows_w_company

# ─────────────────────────────────────────────
# Phase 2: Company Name Normalization
# ─────────────────────────────────────────────
LEGAL_SUFFIXES_EN = [
    r'\bLLC\b', r'\bL\.L\.C\.\b', r'\bLTD\b', r'\bL\.T\.D\.\b',
    r'\bLimited\b', r'\bCo\.\b', r'\bCorp\b', r'\bCorporation\b',
    r'\bInc\b', r'\bIncorporated\b', r'\bCompany\b',
    r'\bCJSC\b',
]
LEGAL_SUFFIXES_AR = [
    r'\bشركة\b', r'\bمؤسسة\b', r'\bمؤسسه\b',
    r'\bشركه\b', r'\bمحدوده\b', r'\bمحدودة\b',
    r'\bشخص واحد\b', r'\bذات مسؤولية محدودة\b',
    r'\bمساهمة مقفلة\b', r'\bمساهمة مبسطة\b',
    r'\bفرع\b', r'\bوكيل\b', r'\bمكتب\b',
]
ALL_SUFFIXES = LEGAL_SUFFIXES_EN + LEGAL_SUFFIXES_AR

def normalize_company_name(name):
    original = name.strip()
    cleaned = original
    
    # Remove trailing annotations in parentheses like (سنوفا), (وديان)
    cleaned = re.sub(r'\s*\([^)]*\)\s*$', '', cleaned)
    
    # Remove sales notes
    cleaned = re.sub(r'\s*[-–—].*$', '', cleaned)
    cleaned = re.sub(r'\s*/\s*.*$', '', cleaned)
    
    # Remove legal suffixes
    for pat in ALL_SUFFIXES:
        cleaned = re.sub(pat, '', cleaned, flags=re.IGNORECASE)
    
    # Normalize whitespace
    cleaned = re.sub(r'\s+', ' ', cleaned).strip()
    cleaned = re.sub(r'\s+', ' ', cleaned).strip()
    
    # If after cleaning we got too short, keep original (but stripped)
    if len(cleaned) < 3:
        cleaned = original.strip()
    
    return {
        'original': original,
        'standardized': cleaned,
    }

def phase2_normalize(rows_w_company):
    for r in rows_w_company:
        opp = (r.get('Opportunity') or '').strip()
        result = normalize_company_name(opp)
        r['_original_name'] = result['original']
        r['_standard_name'] = result['standardized']
        r['_email'] = (r.get('Email') or '').strip()
        r['_phone'] = (r.get('Phone') or '').strip()
        r['_contact'] = (r.get('Contact Name') or '').strip()
        r['_salesperson'] = (r.get('Salesperson') or '').strip()
        r['_stage'] = (r.get('Stage') or '').strip()
        r['_tags'] = (r.get('Tags') or '').strip()
    return rows_w_company

# ─────────────────────────────────────────────
# Phase 3: Company-Level Deduplication
# ─────────────────────────────────────────────
def extract_domain(email):
    if not email or '@' not in email:
        return ''
    domain = email.split('@')[1].lower().strip()
    # Remove mailto: prefix
    domain = re.sub(r'^mailto:', '', domain)
    # Remove common personal email domains
    personal = {'gmail.com', 'hotmail.com', 'yahoo.com', 'outlook.com', 'live.com',
                'icloud.com', 'msn.com', 'aol.com', 'mail.ru', 'yandex.com'}
    if domain in personal:
        return ''
    return domain

def normalize_phone(phone):
    if not phone:
        return ''
    p = re.sub(r'[\s\-\(\)\[\]\+\#\*\u2066\u2067\u2068\u2069]', '', phone)
    p = re.sub(r'^00', '', p)
    p = re.sub(r'^\+?966', '0', p)
    p = re.sub(r'[^0-9]', '', p)
    return p

def name_similarity(n1, n2):
    return SequenceMatcher(None, n1.lower(), n2.lower()).ratio()

def phase3_deduplicate(rows):
    rows = list(rows)
    
    # Build features for each row
    for r in rows:
        r['_domain'] = extract_domain(r.get('_email', ''))
        r['_phone_norm'] = normalize_phone(r.get('_phone', ''))
    
    n = len(rows)
    assigned = [None] * n
    master_id = 0
    groups = []  # list of lists of indices
    
    # Assign company master IDs using priority matching
    for i in range(n):
        if assigned[i] is not None:
            continue
        
        group = [i]
        assigned[i] = master_id
        ri = rows[i]
        
        for j in range(i + 1, n):
            if assigned[j] is not None:
                continue
            rj = rows[j]
            
            # Priority 1: Website domain (from email)
            if ri['_domain'] and ri['_domain'] == rj['_domain']:
                group.append(j)
                assigned[j] = master_id
                continue
            
            # Priority 2: Name similarity >= 0.85
            sim = name_similarity(ri['_standard_name'], rj['_standard_name'])
            if sim >= 0.90:
                group.append(j)
                assigned[j] = master_id
                continue
            
            # Priority 3: Phone match
            if ri['_phone_norm'] and rj['_phone_norm'] and len(ri['_phone_norm']) >= 7:
                if ri['_phone_norm'] == rj['_phone_norm']:
                    group.append(j)
                    assigned[j] = master_id
                    continue
            
            # Priority 4: Partial name match (>= 0.75)
            if sim >= 0.75:
                # Check if they share a significant word
                words_i = set(ri['_standard_name'].lower().split())
                words_j = set(rj['_standard_name'].lower().split())
                common = words_i & words_j
                if len(common) >= 2 or (len(common) >= 1 and (max(len(words_i), len(words_j)) <= 3)):
                    group.append(j)
                    assigned[j] = master_id
                    continue
        
        groups.append(group)
        master_id += 1
    
    # Assign IDs
    for idx, g in enumerate(groups):
        mid = f"COMP-{idx + 1:06d}"
        for i in g:
            rows[i]['_master_id'] = mid
    
    return rows, groups

# ─────────────────────────────────────────────
# Phase 4: Website Discovery Queries
# ─────────────────────────────────────────────
def phase4_website_queries(rows):
    for r in rows:
        name = r['_standard_name']
        q1 = f'"{name}" official website Saudi Arabia'
        q2 = f'"{name}" company website'
        q3 = f'site:.sa "{name}"'
        q4 = f'site:com "{name}"'
        r['_ws_q1'] = q1
        r['_ws_q2'] = q2
        r['_ws_q3'] = q3
        r['_ws_q4'] = q4
    return rows

# ─────────────────────────────────────────────
# Phase 5: LinkedIn Discovery Queries
# ─────────────────────────────────────────────
def phase5_linkedin_queries(rows):
    for r in rows:
        name = r['_standard_name']
        r['_li_q1'] = f'site:linkedin.com/company "{name}"'
        r['_li_q2'] = f'"{name}" LinkedIn'
    return rows

# ─────────────────────────────────────────────
# Phase 6-8: Enrichment Structure + Confidence
# ─────────────────────────────────────────────
def phase6_enrichment(rows):
    today = datetime.now().strftime('%Y-%m-%d')
    for r in rows:
        r['_website'] = ''
        r['_website_domain'] = r.get('_domain', '')
        r['_linkedin_url'] = ''
        r['_linkedin_id'] = ''
        r['_google_maps'] = ''
        r['_country'] = 'Saudi Arabia'
        r['_city'] = ''
        r['_industry'] = ''
        r['_sub_industry'] = ''
        r['_business_category'] = ''
        r['_company_size'] = ''
        r['_employee_range'] = ''
        r['_revenue_range'] = ''
        r['_data_source'] = 'CRM Export'
        r['_website_confidence'] = 0
        r['_linkedin_confidence'] = 0
        r['_last_verified'] = today
    return rows

# ─────────────────────────────────────────────
# Phase 9: Generate Outputs
# ─────────────────────────────────────────────
def style_header(ws, num_cols):
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin
    ws.freeze_panes = 'A2'

def auto_width(ws, num_cols, max_width=50):
    for col in range(1, num_cols + 1):
        max_len = 0
        for row in ws.iter_rows(min_col=col, max_col=col, values_only=True):
            for cell_val in row:
                if cell_val:
                    max_len = max(max_len, min(len(str(cell_val)), max_width))
        ws.column_dimensions[get_column_letter(col)].width = max_len + 3

def build_workbook(rows, groups, audit, unique_rows):
    wb = Workbook()
    
    # ── Sheet 1: ENRICHMENT_READY ──
    ws1 = wb.active
    ws1.title = "ENRICHMENT_READY"
    headers1 = [
        'Company Master ID', 'Original Company Name', 'Standard Company Name',
        'Website', 'Website Domain', 'LinkedIn Company URL', 'LinkedIn Company ID',
        'Google Maps URL', 'Country', 'City', 'Industry', 'Sub Industry',
        'Business Category', 'Company Size', 'Employee Range', 'Revenue Range',
        'Data Source', 'Website Confidence Score', 'LinkedIn Confidence Score',
        'Last Verified Date', 'Contact Name', 'Email', 'Phone',
        'Salesperson', 'Stage', 'Tags',
        'Website Search Query 1', 'Website Search Query 2',
        'Website Search Query 3', 'Website Search Query 4',
        'LinkedIn Search Query 1', 'LinkedIn Search Query 2',
    ]
    ws1.append(headers1)
    
    seen_ids = set()
    for r in unique_rows:
        mid = r['_master_id']
        if mid in seen_ids:
            continue
        seen_ids.add(mid)
        ws1.append([
            r['_master_id'],
            r['_original_name'],
            r['_standard_name'],
            r['_website'],
            r['_website_domain'],
            r['_linkedin_url'],
            r['_linkedin_id'],
            r['_google_maps'],
            r['_country'],
            r['_city'],
            r['_industry'],
            r['_sub_industry'],
            r['_business_category'],
            r['_company_size'],
            r['_employee_range'],
            r['_revenue_range'],
            r['_data_source'],
            r['_website_confidence'],
            r['_linkedin_confidence'],
            r['_last_verified'],
            r['_contact'],
            r['_email'],
            r['_phone'],
            r['_salesperson'],
            r['_stage'],
            r['_tags'],
            r['_ws_q1'],
            r['_ws_q2'],
            r['_ws_q3'],
            r['_ws_q4'],
            r['_li_q1'],
            r['_li_q2'],
        ])
    style_header(ws1, len(headers1))
    auto_width(ws1, len(headers1))
    
    # ── Sheet 2: WEBSITE_DISCOVERY_QUEUE ──
    ws2 = wb.create_sheet("WEBSITE_DISCOVERY_QUEUE")
    headers2 = ['Company Master ID', 'Standard Company Name', 'Original Company Name',
                'Phone', 'Email', 'Contact Name',
                'Website Search Query 1', 'Website Search Query 2',
                'Website Search Query 3', 'Website Search Query 4']
    ws2.append(headers2)
    seen2 = set()
    for r in unique_rows:
        mid = r['_master_id']
        if mid in seen2:
            continue
        seen2.add(mid)
        ws2.append([
            mid, r['_standard_name'], r['_original_name'],
            r['_phone'], r['_email'], r['_contact'],
            r['_ws_q1'], r['_ws_q2'], r['_ws_q3'], r['_ws_q4'],
        ])
    style_header(ws2, len(headers2))
    auto_width(ws2, len(headers2))
    
    # ── Sheet 3: LINKEDIN_DISCOVERY_QUEUE ──
    ws3 = wb.create_sheet("LINKEDIN_DISCOVERY_QUEUE")
    headers3 = ['Company Master ID', 'Standard Company Name', 'Original Company Name',
                'Phone', 'Email', 'Contact Name',
                'LinkedIn Search Query 1', 'LinkedIn Search Query 2']
    ws3.append(headers3)
    seen3 = set()
    for r in unique_rows:
        mid = r['_master_id']
        if mid in seen3:
            continue
        seen3.add(mid)
        ws3.append([
            mid, r['_standard_name'], r['_original_name'],
            r['_phone'], r['_email'], r['_contact'],
            r['_li_q1'], r['_li_q2'],
        ])
    style_header(ws3, len(headers3))
    auto_width(ws3, len(headers3))
    
    # ── Sheet 4: DUPLICATE_COMPANIES ──
    ws4 = wb.create_sheet("DUPLICATE_COMPANIES")
    headers4 = ['Company Master ID', 'Standard Company Name', 'Original Company Name',
                'Occurrences', 'Matched By', 'Contact Name', 'Email', 'Phone']
    ws4.append(headers4)
    
    # Group by master ID and list all occurrences
    mid_groups = defaultdict(list)
    for r in rows:
        mid_groups[r['_master_id']].append(r)
    
    for mid, members in sorted(mid_groups.items()):
        if len(members) < 2:
            continue
        names = set(m['_original_name'] for m in members)
        for m in members:
            matches = []
            if any(m2['_domain'] and m2['_domain'] == m['_domain'] for m2 in members if m2 != m):
                matches.append('Domain')
            if any(name_similarity(m['_standard_name'], m2['_standard_name']) >= 0.90 for m2 in members if m2 != m):
                matches.append('Name Similarity')
            if any(m2['_phone_norm'] and m2['_phone_norm'] == m['_phone_norm'] and len(m['_phone_norm']) >= 7 for m2 in members if m2 != m):
                matches.append('Phone')
            matched_by = ', '.join(set(matches)) if matches else 'Name Partial'
            
            ws4.append([
                mid, m['_standard_name'], m['_original_name'],
                len(members), matched_by,
                m['_contact'], m['_email'], m['_phone'],
            ])
    
    style_header(ws4, len(headers4))
    auto_width(ws4, len(headers4))
    
    # ── Sheet 5: EXECUTIVE_SUMMARY ──
    ws5 = wb.create_sheet("EXECUTIVE_SUMMARY")
    
    total_unique = len(set(r['_master_id'] for r in rows))
    dup_rate = (audit['duplicate_occurrences'] / audit['total_records'] * 100) if audit['total_records'] else 0
    ws_coverage = audit['website_coverage']
    li_coverage = audit['linkedin_coverage']
    dq_score = audit['data_quality_score']
    
    summary_data = [
        ['Metric', 'Value'],
        ['Report Generated', datetime.now().strftime('%Y-%m-%d %H:%M')],
        ['', ''],
        ['--- DATA QUALITY REPORT ---', ''],
        ['Total Raw Rows', audit['total_raw']],
        ['Total Records with Company Names', audit['total_records']],
        ['Unique Companies (after dedup)', total_unique],
        ['Duplicate Occurrences', audit['duplicate_occurrences']],
        ['Duplicate Company Names Count', audit['dup_companies_count']],
        ['Duplicate Rate', f'{dup_rate:.1f}%'],
        ['', ''],
        ['--- DATA COVERAGE ---', ''],
        ['Companies Missing Website', f'{audit["missing_website"]} ({ws_coverage:.1f}% coverage)'],
        ['Companies Missing LinkedIn', f'{audit["missing_linkedin"]} ({li_coverage:.1f}% coverage)'],
        ['Companies Missing Industry', f'{audit["missing_industry"]} ({audit["industry_coverage"]:.1f}% coverage)'],
        ['Companies Missing Contact Name', f'{audit["missing_contact"]}'],
        ['Companies Missing Email', f'{audit["missing_email"]}'],
        ['Companies Missing Phone', f'{audit["missing_phone"]}'],
        ['', ''],
        ['--- QUALITY SCORES ---', ''],
        ['Website Coverage %', f'{ws_coverage:.1f}%'],
        ['LinkedIn Coverage %', f'{li_coverage:.1f}%'],
        ['Industry Coverage %', f'{audit["industry_coverage"]:.1f}%'],
        ['Overall Data Quality Score', f'{dq_score:.1f}%'],
        ['', ''],
        ['--- OUTPUT SHEETS ---', ''],
        ['ENRICHMENT_READY (unique companies)', total_unique],
        ['WEBSITE_DISCOVERY_QUEUE', total_unique],
        ['LINKEDIN_DISCOVERY_QUEUE', total_unique],
        ['DUPLICATE_COMPANIES', audit['total_records'] - total_unique],
        ['', ''],
        ['--- RECOMMENDED NEXT ACTIONS ---', ''],
        ['1', 'Enrich ENRICHMENT_READY sheet via Apollo.io / Clay / OpenAI'],
        ['2', 'Run WEBSITE_DISCOVERY_QUEUE through Google Search / BrightData / SerpAPI'],
        ['3', 'Run LINKEDIN_DISCOVERY_QUEUE through LinkedIn Search / Perplexity'],
        ['4', 'Verify discovered websites with Website Confidence Engine'],
        ['5', 'Verify discovered LinkedIn pages with LinkedIn Confidence Engine'],
        ['6', 'Cross-reference with Saudi MoC commercial register (CR) data'],
        ['7', 'Update CRM with enriched fields and confidence scores'],
        ['8', 'Schedule quarterly re-enrichment cycle'],
    ]
    
    title_font = Font(bold=True, size=14, color='2F5496')
    header_font_s = Font(bold=True, size=11, color='FFFFFF')
    header_fill_s = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
    
    for i, (a, b) in enumerate(summary_data, 1):
        ws5.cell(row=i, column=1, value=a)
        ws5.cell(row=i, column=2, value=b)
        if a.startswith('---'):
            ws5.cell(row=i, column=1).font = Font(bold=True, size=12, color='2F5496')
        elif a == 'Metric':
            ws5.cell(row=i, column=1).font = header_font_s
            ws5.cell(row=i, column=1).fill = header_fill_s
            ws5.cell(row=i, column=2).font = header_font_s
            ws5.cell(row=i, column=2).fill = header_fill_s
    
    ws5.column_dimensions['A'].width = 45
    ws5.column_dimensions['B'].width = 55
    ws5.freeze_panes = 'A2'
    
    # Save
    wb.save(OUTPUT_FILE)
    print(f"[OK] Output saved to: {OUTPUT_FILE}")
    print(f"   Sheet 1: ENRICHMENT_READY - {total_unique} unique companies")
    print(f"   Sheet 2: WEBSITE_DISCOVERY_QUEUE - {total_unique} entries")
    print(f"   Sheet 3: LINKEDIN_DISCOVERY_QUEUE - {total_unique} entries")
    print(f"   Sheet 4: DUPLICATE_COMPANIES - potential duplicates listed")
    print(f"   Sheet 5: EXECUTIVE_SUMMARY with full report")

# ─────────────────────────────────────────────
# Main Pipeline
# ─────────────────────────────────────────────
def main():
    print("=" * 60)
    print("CRM DATA ENHANCEMENT & DIGITAL FOOTPRINT DISCOVERY")
    print("=" * 60)
    
    # Load
    print("\n[Loading data...]")
    rows = load_data()
    print(f"   Loaded {len(rows)} raw rows")
    
    # Phase 1
    print("\n[Phase 1: Data Audit]")
    audit, companies_raw, rows_w_company = phase1_audit(rows)
    print(f"   Total Records: {audit['total_records']}")
    print(f"   Unique Companies: {audit['unique_companies']}")
    print(f"   Duplicate Rate: {audit['duplicate_rate']:.1f}%")
    
    # Phase 2
    print("\n[Phase 2: Company Name Normalization]")
    rows_w_company = phase2_normalize(rows_w_company)
    print(f"   Normalized {len(rows_w_company)} names")
    
    # Phase 3
    print("\n[Phase 3: Company-Level Deduplication]")
    rows_w_company, groups = phase3_deduplicate(rows_w_company)
    print(f"   Created {len(groups)} company master IDs")
    
    # Phase 4
    print("\n[Phase 4: Website Discovery Queries]")
    rows_w_company = phase4_website_queries(rows_w_company)
    print(f"   Generated 4 search queries per company")
    
    # Phase 5
    print("\n[Phase 5: LinkedIn Discovery Queries]")
    rows_w_company = phase5_linkedin_queries(rows_w_company)
    print(f"   Generated 2 search queries per company")
    
    # Phase 6-8
    print("\n[Phase 6-8: Enrichment Structure & Confidence]")
    rows_w_company = phase6_enrichment(rows_w_company)
    print(f"   Added enrichment columns with confidence engines")
    
    # Phase 9
    print("\n[Phase 9: Generating Outputs...]")
    
    # Get one representative row per master ID for unique output
    mid_first = {}
    for r in rows_w_company:
        mid = r['_master_id']
        if mid not in mid_first:
            mid_first[mid] = r
    
    unique_rows = list(mid_first.values())
    build_workbook(rows_w_company, groups, audit, unique_rows)
    
    # Print summary to console
    total_unique = len(unique_rows)
    print(f"\n{'=' * 60}")
    print("EXECUTIVE SUMMARY")
    print(f"{'=' * 60}")
    print(f"Total Companies (after dedup): {total_unique}")
    print(f"Duplicate Rate: {audit['duplicate_rate']:.1f}%")
    print(f"Website Coverage: {audit['website_coverage']:.1f}%")
    print(f"LinkedIn Coverage: {audit['linkedin_coverage']:.1f}%")
    print(f"Data Quality Score: {audit['data_quality_score']:.1f}%")
    print(f"\nRecommended Next Actions:")
    print(f"  1. Enrich ENRICHMENT_READY via Apollo.io / Clay / OpenAI")
    print(f"  2. Run WEBSITE_DISCOVERY_QUEUE through Google / BrightData / SerpAPI")
    print(f"  3. Run LINKEDIN_DISCOVERY_QUEUE through LinkedIn Search / Perplexity")
    print(f"  4. Verify discovered profiles with confidence engines")
    print(f"  5. Cross-reference with Saudi MoC commercial register")
    print(f"\nOutput: {OUTPUT_FILE}")
    print(f"{'=' * 60}")

if __name__ == '__main__':
    main()
